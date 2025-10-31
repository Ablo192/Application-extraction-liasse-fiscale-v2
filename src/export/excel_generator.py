"""
Module de génération de fichiers Excel pour les liasses fiscales.

Ce module crée un fichier Excel avec 2 onglets :
- Onglet 1 : Données fiscales brutes (actif, passif, CR, échéances, affectation)
- Onglet 2 : Analyse financière avec ratios calculés
"""

import openpyxl
import openpyxl.styles
import openpyxl.utils
from src.analysis.ratios import calculer_ratios_financiers


# Structure de l'onglet Analyse Financière
STRUCTURE_ANALYSE = [
    ("=== ACTIVITÉ & RENTABILITÉ (K€) ===", None, None),
    ("Durée (en mois)", "duree_mois", None),
    ("CA", "ca", None),
    ("Production stockée + immobilisée", "prod_stockee_immo", None),
    ("Production globale", "prod_globale", None),
    ("AACE", "aace", None),
    ("Dont sous-traitance", "sous_traitance", None),
    ("% Production globale", "pct_sous_traitance", "%"),
    ("Production interne", "prod_interne", None),
    ("Consommation de matières premières et marchandises", "conso_matieres", None),
    ("% production interne", "pct_conso_matieres", "%"),
    ("Charge de personnel", "charge_personnel", None),
    ("Intérim", "interim", None),
    ("% production interne", "pct_charge_personnel", "%"),
    ("EBE", "ebe", None),
    ("% production interne", "pct_ebe", "%"),
    ("Résultat d'exploitation", "resultat_exploitation", None),
    ("% production interne", "pct_resultat_exploitation", "%"),
    ("Charges financières", "charges_financieres", None),
    ("%EBE", "pct_charges_financieres", "%"),
    ("Résultat exceptionnel", "resultat_exceptionnel", None),
    ("Résultat net", "resultat_net", None),
    ("%PI", "pct_resultat_net", "%"),
    ("CAF (y.c crédit bail)", "caf", None),
    ("", None, None),  # Ligne vide
    ("=== BILAN (K€) ===", None, None),
    ("Durée (en mois)", "duree_mois", None),
    ("Non-valeurs", "non_valeurs", None),
    ("Total bilan", "total_bilan", None),
    ("Capitaux propres", "capitaux_propres", None),
    ("Solvabilité (%)", "solvabilite", "%"),
    ("Couverture de l'activité (%)", "couverture_activite", "%"),
    ("Dette brute", "dette_brute", None),
    ("Gearing brut", "gearing_brut", None),
    ("Leverage brut", "leverage_brut", None),
    ("Dont MLT", "dont_mlt", None),
    ("Dont crédit bail", "dont_cb", None),
    ("Capacité de remboursement", "capacite_remboursement", None),
    ("Annuités à venir", "annuites", None),
    ("Couverture des annuités à venir avec la CAF", "couverture_annuites", None),
    ("Dette nette", "dette_nette", None),
    ("Gearing net", "gearing_net", None),
    ("Leverage net", "leverage_net", None),
    ("C/C Actif", "cc_actif", None),
    ("C/C Passif", "cc_passif", None),
    ("Dividendes", "dividendes", None),
    ("", None, None),  # Ligne vide
    ("=== CYCLE D'EXPLOITATION (K€) ===", None, None),
    ("Durée (en mois)", "duree_mois", None),
    ("FRNG", "frng", None),
    ("BFR", "bfr", None),
    ("Dont BFRE", "bfre", None),
    ("Nb jours", "nb_jours_bfre", "jours"),
    ("Dont Stocks", "stocks", None),
    ("Nb jours", "nb_jours_stocks", "jours"),
    ("Dont créances clients", "creances_clients", None),
    ("Nb jours", "nb_jours_creances", "jours"),
    ("% créances douteuses", "pct_creances_douteuses", "%"),
    ("% créances douteuses provisionnées", "pct_creances_douteuses_prov", "%"),
    ("Dont dettes fournisseurs", "dettes_fournisseurs", None),
    ("Nb jours", "nb_jours_fournisseurs", "jours"),
    ("Trésorerie nette", "tresorerie_nette", None),
]


def creer_fichier_excel(donnees_par_annee, nom_fichier):
    """Crée le fichier Excel avec 2 onglets : Données Fiscales + Analyse Financière.

    Args:
        donnees_par_annee: Dict avec structure {
            'annee1': {'actif': [...], 'passif': [...], 'cr': [...], 'echeances': [...], 'affectation': [...]},
            'annee2': {...}
        }
        nom_fichier: Path du fichier Excel à créer
    """
    print(f"📊 Création du fichier : {nom_fichier.name}")
    wb = openpyxl.Workbook()

    # Trier les années chronologiquement
    annees_triees = sorted(donnees_par_annee.keys())
    print(f"   📅 Années détectées : {', '.join(annees_triees)}")

    # ========================================
    # ONGLET 1: DONNÉES FISCALES
    # ========================================
    ws_donnees = wb.active
    ws_donnees.title = "Données Fiscales"
    _creer_onglet_donnees(ws_donnees, donnees_par_annee, annees_triees)

    # ========================================
    # ONGLET 2: ANALYSE FINANCIÈRE
    # ========================================
    print("   📊 Calcul des ratios financiers...")
    ratios_par_annee = calculer_ratios_financiers(donnees_par_annee)

    ws_analyse = wb.create_sheet("Analyse Financière")
    _creer_onglet_analyse(ws_analyse, ratios_par_annee, annees_triees)

    # Sauvegarder
    wb.save(nom_fichier)
    print(f"✅ Fichier créé avec 2 onglets (Données + Analyse) et {len(annees_triees)} année(s)\n")


def _creer_onglet_donnees(ws, donnees_par_annee, annees_triees):
    """Crée l'onglet Données Fiscales avec les 5 sections.

    Args:
        ws: Worksheet openpyxl
        donnees_par_annee: Dict des données extraites
        annees_triees: Liste des années triées
    """
    # En-têtes
    ws['A1'] = "Catégorie"
    ws['A1'].font = openpyxl.styles.Font(bold=True, size=12)
    ws['B1'] = "Libellé"
    ws['B1'].font = openpyxl.styles.Font(bold=True, size=12)

    for col_idx, annee in enumerate(annees_triees, start=3):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = annee
        cell.font = openpyxl.styles.Font(bold=True, size=12)

    current_row = 2

    # Section 1: BILAN ACTIF
    current_row = _ajouter_section(ws, donnees_par_annee, annees_triees, current_row, "actif", "BILAN ACTIF")
    current_row += 1

    # Section 2: BILAN PASSIF
    current_row = _ajouter_section(ws, donnees_par_annee, annees_triees, current_row, "passif", "BILAN PASSIF")
    current_row += 1

    # Section 3: COMPTE DE RÉSULTAT
    current_row = _ajouter_section(ws, donnees_par_annee, annees_triees, current_row, "cr", "COMPTE RÉSULTAT", bold_keywords=["TOTAL", "RÉSULTAT", "CHIFFRE D'AFFAIRES", "BÉNÉFICE", "PERTE"])
    current_row += 1

    # Section 4: ÉTAT DES ÉCHÉANCES
    current_row = _ajouter_section(ws, donnees_par_annee, annees_triees, current_row, "echeances", "ÉCHÉANCES")
    current_row += 1

    # Section 5: AFFECTATION & RENSEIGNEMENTS
    current_row = _ajouter_section(ws, donnees_par_annee, annees_triees, current_row, "affectation", "AFFECTATION")

    # Ajuster largeur colonnes
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 60
    for col_idx in range(3, len(annees_triees) + 3):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15

    # Figer en-têtes
    ws.freeze_panes = 'C2'


def _ajouter_section(ws, donnees_par_annee, annees_triees, start_row, section_key, section_name, bold_keywords=None):
    """Ajoute une section de données au worksheet.

    Args:
        ws: Worksheet openpyxl
        donnees_par_annee: Dict des données extraites
        annees_triees: Liste des années triées
        start_row: Numéro de ligne de départ
        section_key: Clé de la section ('actif', 'passif', 'cr', etc.)
        section_name: Nom affiché dans la colonne Catégorie
        bold_keywords: Liste de mots-clés pour mettre en gras (optionnel)

    Returns:
        int: Numéro de la prochaine ligne disponible
    """
    if bold_keywords is None:
        bold_keywords = ["TOTAL"]

    current_row = start_row

    if annees_triees:
        premiere_annee = annees_triees[0]
        donnees_section = donnees_par_annee[premiere_annee].get(section_key, [])

        for libelle, _ in donnees_section:
            ws[f'A{current_row}'] = section_name
            ws[f'B{current_row}'] = libelle

            # Mettre en gras si nécessaire
            should_bold = any(keyword in libelle.upper() for keyword in bold_keywords)
            if should_bold:
                ws[f'A{current_row}'].font = openpyxl.styles.Font(bold=True)
                ws[f'B{current_row}'].font = openpyxl.styles.Font(bold=True)

            # Remplir les montants pour chaque année
            for col_idx, annee in enumerate(annees_triees, start=3):
                donnees_annee = donnees_par_annee[annee].get(section_key, [])
                montant = next((m for l, m in donnees_annee if l == libelle), 0)

                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = montant
                cell.number_format = '#,##0.00'

                if should_bold:
                    cell.font = openpyxl.styles.Font(bold=True)

            current_row += 1

    return current_row


def _creer_onglet_analyse(ws, ratios_par_annee, annees_triees):
    """Crée l'onglet Analyse Financière avec les ratios.

    Args:
        ws: Worksheet openpyxl
        ratios_par_annee: Dict des ratios calculés
        annees_triees: Liste des années triées
    """
    # En-têtes
    ws['A1'] = "Indicateur"
    ws['A1'].font = openpyxl.styles.Font(bold=True, size=12)

    for col_idx, annee in enumerate(annees_triees, start=2):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = annee
        cell.font = openpyxl.styles.Font(bold=True, size=12)

    current_row = 2

    # Remplir les données selon la structure
    for libelle, cle_ratio, format_type in STRUCTURE_ANALYSE:
        ws[f'A{current_row}'] = libelle

        # Titres de sections en gras
        if libelle.startswith("==="):
            ws[f'A{current_row}'].font = openpyxl.styles.Font(bold=True, size=12)
            current_row += 1
            continue

        # Ligne vide
        if not libelle:
            current_row += 1
            continue

        # Remplir les valeurs pour chaque année
        for col_idx, annee in enumerate(annees_triees, start=2):
            if cle_ratio and annee in ratios_par_annee:
                valeur = ratios_par_annee[annee].get(cle_ratio, 0)

                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = valeur

                # Format selon le type
                if format_type == "%":
                    cell.number_format = '0.00"%"'
                elif format_type == "jours":
                    cell.number_format = '0.0'
                else:
                    cell.number_format = '#,##0.00'

        current_row += 1

    # Ajuster largeur colonnes
    ws.column_dimensions['A'].width = 60
    for col_idx in range(2, len(annees_triees) + 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15

    # Figer en-têtes
    ws.freeze_panes = 'B2'
