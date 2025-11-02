#!/usr/bin/env python3
"""
Application d'extraction de liasses fiscales françaises depuis des PDFs avec Camelot.

Ce script utilise Camelot (au lieu de pdfplumber) pour une extraction plus robuste
des tableaux. Camelot offre deux modes:
- 'lattice': pour les tableaux avec des lignes visibles (par défaut)
- 'stream': pour les tableaux sans lignes (fallback)

Usage:
    python main_camelot.py

Le script traite tous les PDFs du dossier 'liasses/' et génère un fichier Excel
dans le dossier 'resultats/'.
"""

import camelot
import pdfplumber
from pathlib import Path
import pandas as pd


# ============================================================
# UTILITAIRES PDF
# ============================================================

def extraire_annee_fiscale(chemin_pdf):
    """Extrait l'année fiscale du PDF (utilise pdfplumber pour le texte)."""
    try:
        with pdfplumber.open(chemin_pdf) as pdf:
            for page in pdf.pages[:3]:  # Rechercher dans les 3 premières pages
                texte = page.extract_text()
                if not texte:
                    continue

                # Chercher des patterns d'année
                import re
                patterns = [
                    r'Exercice.*?(\d{4})',
                    r'du.*?(\d{2})/(\d{2})/(\d{4}).*?au.*?(\d{2})/(\d{2})/(\d{4})',
                    r'(\d{4})',
                ]

                for pattern in patterns:
                    matches = re.findall(pattern, texte)
                    if matches:
                        if isinstance(matches[0], tuple):
                            # Pour les dates complètes, prendre la dernière année
                            annees = [m for m in matches[0] if len(m) == 4]
                            if annees:
                                return annees[-1]
                        elif len(matches[0]) == 4:
                            return matches[0]
    except Exception as e:
        print(f"⚠️  Erreur lors de l'extraction de l'année : {e}")

    return None


def trouver_page_contenant(chemin_pdf, mots_cles):
    """Trouve le numéro de page contenant les mots-clés (utilise pdfplumber)."""
    try:
        with pdfplumber.open(chemin_pdf) as pdf:
            for idx, page in enumerate(pdf.pages):
                texte = page.extract_text()
                if texte and any(mot.upper() in texte.upper() for mot in mots_cles):
                    return idx
    except Exception as e:
        print(f"⚠️  Erreur lors de la recherche de page : {e}")

    return -1


def extraire_tables_camelot(chemin_pdf, page_num, mode='lattice'):
    """
    Extrait les tables d'une page avec Camelot.

    Args:
        chemin_pdf: Path du PDF
        page_num: Numéro de page (1-indexed pour Camelot)
        mode: 'lattice' (défaut) ou 'stream'

    Returns:
        list de DataFrames pandas
    """
    try:
        print(f"   📊 Mode Camelot: {mode}")
        tables = camelot.read_pdf(
            str(chemin_pdf),
            pages=str(page_num),
            flavor=mode,
            suppress_stdout=True
        )

        if tables:
            print(f"   ✓ {len(tables)} tableau(x) extrait(s)")
            # Afficher quelques infos sur la qualité
            for i, table in enumerate(tables):
                if hasattr(table, 'parsing_report'):
                    accuracy = table.parsing_report.get('accuracy', 0)
                    print(f"      Table {i+1}: précision = {accuracy:.1f}%")

            return [table.df for table in tables]
        else:
            print(f"   ❌ Aucun tableau trouvé")
            return []

    except Exception as e:
        print(f"   ❌ Erreur Camelot ({mode}): {e}")
        return []


# ============================================================
# EXTRACTORS ADAPTÉS POUR CAMELOT
# ============================================================

def nettoyer_valeur(valeur):
    """Nettoie une valeur extraite."""
    if pd.isna(valeur) or valeur == '':
        return ''

    valeur = str(valeur).strip()
    # Supprimer les espaces dans les nombres
    if valeur.replace(' ', '').replace('-', '').replace(',', '').isdigit():
        valeur = valeur.replace(' ', '')

    return valeur


def extraire_bilan_actif(df):
    """
    Extrait les données du Bilan Actif depuis un DataFrame Camelot.
    Format attendu: LABEL | Brut | Amort./Prov. | Net N | Net N-1
    """
    donnees = []

    print(f"   📋 DataFrame shape: {df.shape}")

    # Chercher la ligne d'en-tête
    header_row = None
    for idx, row in df.iterrows():
        row_text = ' '.join(str(cell).upper() for cell in row)
        if 'BRUT' in row_text and 'NET' in row_text:
            header_row = idx
            print(f"   ✓ En-tête trouvé à la ligne {idx}")
            break

    if header_row is None:
        print("   ⚠️  En-tête non trouvé, utilisation de la première ligne")
        header_row = 0

    # Extraire les données ligne par ligne
    for idx in range(header_row + 1, len(df)):
        row = df.iloc[idx]

        # Vérifier qu'on a assez de colonnes
        if len(row) < 5:
            continue

        label = nettoyer_valeur(row.iloc[0])

        # Ignorer les lignes vides ou en-têtes
        if not label or label.upper() in ['ACTIF', 'BRUT', 'TOTAL']:
            continue

        # Extraire les valeurs
        brut = nettoyer_valeur(row.iloc[1])
        amort = nettoyer_valeur(row.iloc[2])
        net_n = nettoyer_valeur(row.iloc[3])
        net_n1 = nettoyer_valeur(row.iloc[4]) if len(row) > 4 else ''

        donnees.append({
            'Ligne': label,
            'Brut': brut,
            'Amortissement/Provision': amort,
            'Net (N)': net_n,
            'Net (N-1)': net_n1
        })

    return donnees


def extraire_bilan_passif(df):
    """
    Extrait les données du Bilan Passif depuis un DataFrame Camelot.
    Format attendu: LABEL | Net N | Net N-1
    """
    donnees = []

    print(f"   📋 DataFrame shape: {df.shape}")

    # Chercher la ligne d'en-tête
    header_row = None
    for idx, row in df.iterrows():
        row_text = ' '.join(str(cell).upper() for cell in row)
        if 'PASSIF' in row_text or 'CAPITAL' in row_text:
            header_row = idx
            print(f"   ✓ En-tête trouvé à la ligne {idx}")
            break

    if header_row is None:
        print("   ⚠️  En-tête non trouvé, utilisation de la première ligne")
        header_row = 0

    # Extraire les données ligne par ligne
    for idx in range(header_row + 1, len(df)):
        row = df.iloc[idx]

        # Vérifier qu'on a assez de colonnes
        if len(row) < 3:
            continue

        label = nettoyer_valeur(row.iloc[0])

        # Ignorer les lignes vides ou en-têtes
        if not label or label.upper() in ['PASSIF', 'TOTAL']:
            continue

        # Extraire les valeurs
        net_n = nettoyer_valeur(row.iloc[1]) if len(row) > 1 else ''
        net_n1 = nettoyer_valeur(row.iloc[2]) if len(row) > 2 else ''

        donnees.append({
            'Ligne': label,
            'Net (N)': net_n,
            'Net (N-1)': net_n1
        })

    return donnees


def extraire_compte_resultat(chemin_pdf):
    """
    Extrait le Compte de Résultat (formulaires 2052 et 2053).
    Ces formulaires sont souvent sur 2 pages consécutives.
    """
    donnees = []

    # Trouver les pages du CR
    page_2052 = trouver_page_contenant(chemin_pdf, ["COMPTE DE RÉSULTAT", "PRODUITS D'EXPLOITATION"])

    if page_2052 == -1:
        print("   ❌ Page 2052 non trouvée")
        return donnees

    print(f"   ✓ Page 2052 identifiée : {page_2052 + 1}")

    # Extraire les tables des 2 pages (2052 et 2053)
    for offset in [0, 1]:
        page_num = page_2052 + offset + 1  # +1 car Camelot est 1-indexed
        print(f"\n   📄 Extraction page {page_num}...")

        # Essayer d'abord lattice, puis stream
        tables = extraire_tables_camelot(chemin_pdf, page_num, mode='lattice')
        if not tables:
            tables = extraire_tables_camelot(chemin_pdf, page_num, mode='stream')

        for table_idx, df in enumerate(tables):
            print(f"      Table {table_idx + 1}: {df.shape}")

            # Extraire ligne par ligne
            for idx in range(len(df)):
                row = df.iloc[idx]

                if len(row) < 2:
                    continue

                label = nettoyer_valeur(row.iloc[0])

                # Ignorer les lignes vides ou en-têtes
                if not label or len(label) < 3:
                    continue

                # Extraire les valeurs
                exercice_n = nettoyer_valeur(row.iloc[1]) if len(row) > 1 else ''
                exercice_n1 = nettoyer_valeur(row.iloc[2]) if len(row) > 2 else ''

                donnees.append({
                    'Ligne': label,
                    'Exercice N': exercice_n,
                    'Exercice N-1': exercice_n1
                })

    return donnees


def extraire_etat_echeances(df):
    """
    Extrait l'État des Échéances depuis un DataFrame Camelot.
    """
    donnees = []

    print(f"   📋 DataFrame shape: {df.shape}")

    # Chercher la ligne d'en-tête
    header_row = None
    for idx, row in df.iterrows():
        row_text = ' '.join(str(cell).upper() for cell in row)
        if 'ÉCHÉANCE' in row_text or 'ECHEANCE' in row_text or 'MOIS' in row_text:
            header_row = idx
            print(f"   ✓ En-tête trouvé à la ligne {idx}")
            break

    if header_row is None:
        print("   ⚠️  En-tête non trouvé, utilisation de la première ligne")
        header_row = 0

    # Extraire les données ligne par ligne
    for idx in range(header_row + 1, len(df)):
        row = df.iloc[idx]

        if len(row) < 2:
            continue

        label = nettoyer_valeur(row.iloc[0])

        # Ignorer les lignes vides
        if not label:
            continue

        # Extraire les valeurs (nombre de colonnes variable)
        valeurs = [nettoyer_valeur(row.iloc[i]) for i in range(1, len(row))]

        donnees.append({
            'Ligne': label,
            'Valeurs': ' | '.join(valeurs)
        })

    return donnees


def extraire_affectation_resultat(df):
    """
    Extrait l'Affectation du Résultat depuis un DataFrame Camelot.
    """
    donnees = []

    print(f"   📋 DataFrame shape: {df.shape}")

    # Extraire toutes les lignes significatives
    for idx in range(len(df)):
        row = df.iloc[idx]

        if len(row) < 2:
            continue

        label = nettoyer_valeur(row.iloc[0])

        # Ignorer les lignes vides ou titres
        if not label or label.upper() in ['AFFECTATION', 'RESULTAT']:
            continue

        # Extraire la valeur
        valeur = nettoyer_valeur(row.iloc[1]) if len(row) > 1 else ''

        donnees.append({
            'Ligne': label,
            'Montant': valeur
        })

    return donnees


# ============================================================
# GÉNÉRATION EXCEL
# ============================================================

def creer_fichier_excel(donnees_par_annee, nom_fichier):
    """
    Génère un fichier Excel avec toutes les données extraites.

    Args:
        donnees_par_annee: dict {annee: {actif, passif, cr, echeances, affectation}}
        nom_fichier: Path du fichier Excel à créer
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    wb.remove(wb.active)  # Supprimer la feuille par défaut

    # Style pour les en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for annee, donnees in sorted(donnees_par_annee.items()):
        # Créer une feuille par année
        ws = wb.create_sheet(title=f"Année {annee}")
        row = 1

        # BILAN ACTIF
        if donnees.get('actif'):
            ws.cell(row, 1, "BILAN ACTIF (2050)").font = Font(bold=True, size=14)
            row += 1

            # En-têtes
            headers = ['Ligne', 'Brut', 'Amort./Prov.', 'Net (N)', 'Net (N-1)']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row, col, header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            row += 1

            # Données
            for item in donnees['actif']:
                ws.cell(row, 1, item.get('Ligne', ''))
                ws.cell(row, 2, item.get('Brut', ''))
                ws.cell(row, 3, item.get('Amortissement/Provision', ''))
                ws.cell(row, 4, item.get('Net (N)', ''))
                ws.cell(row, 5, item.get('Net (N-1)', ''))
                row += 1

            row += 2

        # BILAN PASSIF
        if donnees.get('passif'):
            ws.cell(row, 1, "BILAN PASSIF (2051)").font = Font(bold=True, size=14)
            row += 1

            headers = ['Ligne', 'Net (N)', 'Net (N-1)']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row, col, header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            row += 1

            for item in donnees['passif']:
                ws.cell(row, 1, item.get('Ligne', ''))
                ws.cell(row, 2, item.get('Net (N)', ''))
                ws.cell(row, 3, item.get('Net (N-1)', ''))
                row += 1

            row += 2

        # COMPTE DE RÉSULTAT
        if donnees.get('cr'):
            ws.cell(row, 1, "COMPTE DE RÉSULTAT (2052-2053)").font = Font(bold=True, size=14)
            row += 1

            headers = ['Ligne', 'Exercice N', 'Exercice N-1']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row, col, header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            row += 1

            for item in donnees['cr']:
                ws.cell(row, 1, item.get('Ligne', ''))
                ws.cell(row, 2, item.get('Exercice N', ''))
                ws.cell(row, 3, item.get('Exercice N-1', ''))
                row += 1

            row += 2

        # ÉTAT DES ÉCHÉANCES
        if donnees.get('echeances'):
            ws.cell(row, 1, "ÉTAT DES ÉCHÉANCES (2057)").font = Font(bold=True, size=14)
            row += 1

            headers = ['Ligne', 'Valeurs']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row, col, header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            row += 1

            for item in donnees['echeances']:
                ws.cell(row, 1, item.get('Ligne', ''))
                ws.cell(row, 2, item.get('Valeurs', ''))
                row += 1

            row += 2

        # AFFECTATION DU RÉSULTAT
        if donnees.get('affectation'):
            ws.cell(row, 1, "AFFECTATION DU RÉSULTAT (2058-C)").font = Font(bold=True, size=14)
            row += 1

            headers = ['Ligne', 'Montant']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row, col, header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            row += 1

            for item in donnees['affectation']:
                ws.cell(row, 1, item.get('Ligne', ''))
                ws.cell(row, 2, item.get('Montant', ''))
                row += 1

        # Ajuster les largeurs de colonnes
        ws.column_dimensions['A'].width = 50
        for col in ['B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 15

    wb.save(nom_fichier)
    print(f"\n✅ Fichier Excel créé : {nom_fichier}")


# ============================================================
# FONCTION PRINCIPALE D'EXTRACTION
# ============================================================

def extraire_un_pdf(chemin_pdf):
    """Extrait les données d'un seul PDF avec Camelot."""
    print(f"\n{'='*80}")
    print(f"📄 Traitement : {chemin_pdf.name}")
    print(f"{'='*80}\n")

    try:
        # Extraire l'année fiscale
        annee = extraire_annee_fiscale(chemin_pdf)
        if annee:
            print(f"📅 Année fiscale détectée : {annee}\n")

        # ============================================================
        # BILAN ACTIF (Formulaire 2050)
        # ============================================================
        print("\n" + "="*80)
        print("📋 BILAN ACTIF (Formulaire 2050)")
        print("="*80)

        print("🔍 Recherche de la page du Bilan Actif...")
        actif_page_index = trouver_page_contenant(chemin_pdf, ["Brut", "Net", "ACTIF"])

        if actif_page_index == -1:
            print("❌ Page non trouvée. Passage au formulaire suivant.")
            donnees_actif = []
        else:
            print(f"   ✓ Page identifiée : {actif_page_index + 1}")

            # Extraire avec Camelot (page_num est 1-indexed)
            tables = extraire_tables_camelot(chemin_pdf, actif_page_index + 1, mode='lattice')
            if not tables:
                # Fallback sur stream
                tables = extraire_tables_camelot(chemin_pdf, actif_page_index + 1, mode='stream')

            if not tables:
                print("❌ Aucun tableau trouvé.")
                donnees_actif = []
            else:
                print("🚀 Traitement des données...")
                donnees_actif = extraire_bilan_actif(tables[0])
                print(f"✅ Actif terminé: {len(donnees_actif)} lignes extraites")

        # ============================================================
        # BILAN PASSIF (Formulaire 2051)
        # ============================================================
        print("\n" + "="*80)
        print("📋 BILAN PASSIF (Formulaire 2051)")
        print("="*80)

        print("🔍 Recherche de la page du Bilan Passif...")
        passif_page_index = trouver_page_contenant(chemin_pdf, ["Capital social ou individuel", "PASSIF"])

        if passif_page_index == -1:
            print("❌ Page non trouvée. Passage au formulaire suivant.")
            donnees_passif = []
        else:
            print(f"   ✓ Page identifiée : {passif_page_index + 1}")

            tables = extraire_tables_camelot(chemin_pdf, passif_page_index + 1, mode='lattice')
            if not tables:
                tables = extraire_tables_camelot(chemin_pdf, passif_page_index + 1, mode='stream')

            if not tables:
                print("❌ Aucun tableau trouvé.")
                donnees_passif = []
            else:
                print("🚀 Traitement des données...")
                donnees_passif = extraire_bilan_passif(tables[0])
                print(f"✅ Passif terminé: {len(donnees_passif)} lignes extraites")

        # ============================================================
        # COMPTE DE RÉSULTAT (Formulaires 2052-2053)
        # ============================================================
        print("\n" + "="*80)
        print("📋 COMPTE DE RÉSULTAT (Formulaires 2052-2053)")
        print("="*80)

        print("🚀 Traitement des données (2 pages)...")
        donnees_cr = extraire_compte_resultat(chemin_pdf)
        print(f"✅ Compte de Résultat terminé: {len(donnees_cr)} lignes extraites")

        # ============================================================
        # ÉTAT DES ÉCHÉANCES (Formulaire 2057)
        # ============================================================
        print("\n" + "="*80)
        print("📋 ÉTAT DES ÉCHÉANCES (Formulaire 2057)")
        print("="*80)

        print("🔍 Recherche de la page de l'État des Échéances...")
        echeances_page_index = trouver_page_contenant(chemin_pdf, ["ÉTAT DES ÉCHÉANCES", "ETAT DES ECHEANCES"])

        if echeances_page_index == -1:
            print("❌ Page non trouvée. Passage au formulaire suivant.")
            donnees_echeances = []
        else:
            print(f"   ✓ Page identifiée : {echeances_page_index + 1}")

            tables = extraire_tables_camelot(chemin_pdf, echeances_page_index + 1, mode='lattice')
            if not tables:
                tables = extraire_tables_camelot(chemin_pdf, echeances_page_index + 1, mode='stream')

            if not tables:
                print("❌ Aucun tableau trouvé.")
                donnees_echeances = []
            else:
                print("🚀 Traitement des données...")
                donnees_echeances = extraire_etat_echeances(tables[0])
                print(f"✅ Échéances terminé: {len(donnees_echeances)} lignes extraites")

        # ============================================================
        # AFFECTATION DU RÉSULTAT (Formulaire 2058-C)
        # ============================================================
        print("\n" + "="*80)
        print("📋 AFFECTATION DU RÉSULTAT (Formulaire 2058-C)")
        print("="*80)

        print("🔍 Recherche de la page de l'Affectation...")
        affectation_page_index = trouver_page_contenant(chemin_pdf, ["AFFECTATION DU RÉSULTAT", "RENSEIGNEMENTS DIVERS"])

        if affectation_page_index == -1:
            print("❌ Page non trouvée. Passage au formulaire suivant.")
            donnees_affectation = []
        else:
            print(f"   ✓ Page identifiée : {affectation_page_index + 1}")

            tables = extraire_tables_camelot(chemin_pdf, affectation_page_index + 1, mode='lattice')
            if not tables:
                tables = extraire_tables_camelot(chemin_pdf, affectation_page_index + 1, mode='stream')

            if not tables:
                print("❌ Aucun tableau trouvé.")
                donnees_affectation = []
            else:
                print("🚀 Traitement des données...")
                donnees_affectation = extraire_affectation_resultat(tables[0])
                print(f"✅ Affectation terminé: {len(donnees_affectation)} lignes extraites")

        return {
            'actif': donnees_actif,
            'passif': donnees_passif,
            'cr': donnees_cr,
            'echeances': donnees_echeances,
            'affectation': donnees_affectation
        }

    except Exception as e:
        print(f"❌ Erreur lors du traitement : {e}")
        import traceback
        traceback.print_exc()
        return None


def main():
    """Point d'entrée principal : traite tous les PDFs du dossier 'liasses/'."""
    print("\n" + "="*80)
    print("🚀 EXTRACTION LIASSE FISCALE - VERSION CAMELOT")
    print("="*80)
    print("\n🔧 Utilisation de Camelot pour une extraction robuste des tableaux")
    print("   Modes: Lattice (tableaux avec lignes) + Stream (fallback)\n")

    dossier_liasses = Path("liasses")
    dossier_resultats = Path("resultats")
    dossier_resultats.mkdir(exist_ok=True)

    fichiers_pdf = list(dossier_liasses.glob("*.pdf"))

    if not fichiers_pdf:
        print("\n❌ Aucun PDF dans 'liasses/'\n")
        return

    print(f"📁 {len(fichiers_pdf)} fichier(s) PDF trouvé(s)")
    print("ℹ️  Chaque PDF sera traité avec détection automatique de l'année")
    print()

    donnees_par_annee = {}

    # Traiter chaque PDF
    for idx, chemin_pdf in enumerate(fichiers_pdf):
        print(f"\n{'='*80}")
        print(f"📄 Fichier {idx + 1}/{len(fichiers_pdf)} : {chemin_pdf.name}")
        print(f"{'='*80}")

        resultats = extraire_un_pdf(chemin_pdf)

        if resultats:
            # Extraire l'année
            annee = extraire_annee_fiscale(chemin_pdf)

            # Si l'année n'est pas détectée, utiliser une attribution automatique
            if not annee:
                annee = str(2023 + idx)
                print(f"\n⚠️ Année non détectée, utilisation de l'année par défaut : {annee}")

            donnees_par_annee[annee] = resultats
            print(f"\n✅ Extraction réussie pour {chemin_pdf.name} (année {annee})")
        else:
            print(f"\n❌ Échec de l'extraction pour {chemin_pdf.name}")

    # Générer le fichier Excel
    if donnees_par_annee:
        print(f"\n{'='*80}")
        print("📊 GÉNÉRATION DU FICHIER EXCEL")
        print(f"{'='*80}\n")

        nom_excel = dossier_resultats / "extraction_multi_annees_camelot.xlsx"
        creer_fichier_excel(donnees_par_annee, nom_excel)

        print("="*80)
        print("✅ EXTRACTION TERMINÉE")
        print("="*80)
        print(f"\n📥 Fichier généré : {nom_excel}")
        print(f"📅 Années extraites : {', '.join(sorted(donnees_par_annee.keys()))}")
        print()
    else:
        print("\n❌ Aucune donnée n'a pu être extraite.\n")


if __name__ == "__main__":
    main()
