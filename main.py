import pdfplumber
import openpyxl
from pathlib import Path

# ============================================
# CONFIGURATION
# ============================================

# --- Dictionnaire pour l'extraction par CODES (ACTIF) ---
CODES_BILAN_ACTIF = {
    "AB": "Frais d'établissement", "CX": "Frais de développement", "AF": "Concessions, brevets et droits similaires",
    "AH": "Fonds commercial", "AJ": "Autres immobilisations incorporelles", "AL": "Avances et acomptes sur immobilisations incorporelles",
    "AN": "Terrains", "AP": "Constructions", "AR": "Installations techniques, matériel et outillage industriels",
    "AT": "Autres immobilisations corporelles", "AV": "Immobilisations en cours", "AX": "Avances et acomptes",
    "CS": "Participations évaluées selon la méthode de mise en équivalence", "CU": "Autres participations",
    "BB": "Créances rattachées à des participations", "BD": "Autres titres immobilisés", "BF": "Prêts",
    "BH": "Autres immobilisations financières", "BJ": "TOTAL ACTIF IMMOBILISÉ", "BL": "Matières premières, approvisionnements",
    "BN": "En cours de production de biens", "BP": "En cours de production de services", "BR": "Produits intermédiaires et finis",
    "BT": "Marchandises", "BV": "Avances et acomptes versés sur commandes", "BX": "Clients et comptes rattachés",
    "BZ": "Autres créances", "CB": "Capital souscrit et appelé, non versé", "CD": "Valeurs mobilières de placement",
    "CF": "Disponibilités", "CH": "Charges constatées d'avance", "CJ": "TOTAL ACTIF CIRCULANT",
    "CM": "Frais d'émission d'emprunt à étaler", "CO": "TOTAL GÉNÉRAL", "CW": "Écarts de conversion actif"
}

# --- Dictionnaire pour l'extraction par CODES (PASSIF) ---
CODES_BILAN_PASSIF = {
    "DA": "Capital social ou individuel",
    "DB": "Primes d'émission, de fusion, d'apport…",
    "DC": "Écarts de réévaluation",
    "DD": "Réserve légale",
    "DE": "Réserves statutaires ou contractuelles",
    "DF": "Réserves réglementées",
    "DG": "Autres réserves",
    "DH": "Report à nouveau",
    "DI": "RÉSULTAT DE L'EXERCICE (bénéfice ou perte)",
    "DJ": "Subventions d'investissement",
    "DK": "Provisions réglementées",
    "DL": "TOTAL (I)",
    "DM": "Produit des émissions de titres participatifs",
    "DN": "Avances conditionnées",
    "DO": "TOTAL (II)",
    "DP": "Provisions pour risques",
    "DQ": "Provisions pour charges",
    "DR": "TOTAL (III)",
    "DS": "Emprunts obligatoires convertibles",
    "DT": "Autres emprunts obligataires",
    "DU": "Emprunts et dettes auprès des établissements de crédit",
    "DV": "Emprunts et dettes financières divers",
    "DW": "Avances et acomptes reçus sur commandes en cours",
    "DX": "Dettes fournisseurs et comptes rattachés",
    "DY": "Dettes fiscales et sociales",
    "DZ": "Dettes sur immobilisations et comptes rattachés",
    "EA": "Autres dettes",
    "EB": "Produits constatés d'avance",
    "EC": "TOTAL (IV)",
    "ED": "Écart de conversion passif",
    "EE": "TOTAL GENERAL (I à V)"
}

# --- Dictionnaire pour l'extraction par CODES (COMPTE DE RÉSULTAT) ---
CODES_COMPTE_RESULTAT = {
    # PRODUITS D'EXPLOITATION
    "FA": "Ventes de marchandises (France)",
    "FB": "Ventes de marchandises (Exportations)",
    "FC": "Ventes de marchandises (TOTAL)",
    "FD": "Production vendue - Biens (France)",
    "FE": "Production vendue - Biens (Exportations)",
    "FF": "Production vendue - Biens (TOTAL)",
    "FG": "Production vendue - Services (France)",
    "FH": "Production vendue - Services (Exportations)",
    "FI": "Production vendue - Services (TOTAL)",
    "FJ": "Chiffres d'affaires nets (France)",
    "FK": "Chiffres d'affaires nets (Exportations)",
    "FL": "Chiffres d'affaires nets (TOTAL)",
    "FM": "Production stockée",
    "FN": "Production immobilisée",
    "FO": "Subventions d'exploitation",
    "FP": "Reprises sur amortissements et provisions, transferts de charges",
    "FQ": "Autres produits",
    "FR": "TOTAL DES PRODUITS D'EXPLOITATION (I)",
    
    # CHARGES D'EXPLOITATION
    "FS": "Achats de marchandises (y compris droits de douane)",
    "FT": "Variation de stocks (marchandises)",
    "FU": "Achats de matières premières et autres approvisionnements",
    "FV": "Variation de stocks (matières premières et approvisionnements)",
    "FW": "Autres achats et charges externes",
    "FX": "Impôts, taxes et versements assimilés",
    "FY": "Salaires et traitements",
    "FZ": "Charges sociales",
    "GA": "Dotations aux amortissements (sur immobilisations)",
    "GB": "Dotations aux provisions (sur immobilisations)",
    "GC": "Dotations aux amortissements (sur actif circulant)",
    "GD": "Dotations aux provisions (sur actif circulant)",
    "GE": "Autres charges",
    "GF": "TOTAL DES CHARGES D'EXPLOITATION (II)",
    
    # RÉSULTAT D'EXPLOITATION
    "GG": "RÉSULTAT D'EXPLOITATION (I - II)",
    
    # OPÉRATIONS EN COMMUN
    "GH": "Bénéfice attribué ou perte transférée (III)",
    "GI": "Perte supportée ou bénéfice transféré (IV)",
    
    # PRODUITS FINANCIERS
    "GJ": "Produits financiers de participations",
    "GK": "Produits des autres valeurs mobilières et créances de l'actif immobilisé",
    "GL": "Autres intérêts et produits assimilés",
    "GM": "Reprises sur provisions et transferts de charges",
    "GN": "Différences positives de change",
    "GO": "Produits nets sur cessions de valeurs mobilières de placement",
    "GP": "TOTAL DES PRODUITS FINANCIERS (V)",
    
    # CHARGES FINANCIÈRES
    "GQ": "Dotations financières aux amortissements et provisions",
    "GR": "Intérêts et charges assimilées",
    "GS": "Différences négatives de change",
    "GT": "Charges nettes sur cessions de valeurs mobilières de placement",
    "GU": "TOTAL DES CHARGES FINANCIÈRES (VI)",
    
    # RÉSULTAT FINANCIER
    "GV": "RÉSULTAT FINANCIER (V - VI)",
    
    # RÉSULTAT COURANT
    "GW": "RÉSULTAT COURANT AVANT IMPÔTS (I - II + III - IV + V - VI)",
    
    # PRODUITS EXCEPTIONNELS
    "HA": "Produits exceptionnels sur opérations de gestion",
    "HB": "Produits exceptionnels sur opérations en capital",
    "HC": "Reprises sur provisions et transferts de charges",
    "HD": "Total des produits exceptionnels (VII)",
    
    # CHARGES EXCEPTIONNELLES
    "HE": "Charges exceptionnelles sur opérations de gestion",
    "HF": "Charges exceptionnelles sur opérations en capital",
    "HG": "Dotations exceptionnelles aux amortissements et provisions",
    "HH": "Total des charges exceptionnelles (VIII)",
    
    # RÉSULTAT EXCEPTIONNEL
    "HI": "RÉSULTAT EXCEPTIONNEL (VII - VIII)",
    
    # PARTICIPATION ET IMPÔTS
    "HJ": "Participation des salariés aux résultats de l'entreprise (IX)",
    "HK": "Impôts sur les bénéfices (X)",
    
    # TOTAUX
    "HL": "TOTAL DES PRODUITS (I + III + V + VII)",
    "HM": "TOTAL DES CHARGES (II + IV + VI + VIII + IX + X)",
    
    # RÉSULTAT NET
    "HN": "BÉNÉFICE OU PERTE (Total des produits - Total des charges)",
    
    # RENVOIS DEMANDÉS
    "HP": "Crédit-bail mobilier",
    "HQ": "Crédit-bail immobilier",
    "A1": "Transfert de charges"
}

# --- Dictionnaire pour l'extraction par CODES (ÉTAT DES ÉCHÉANCES 2057-SD) ---
CODES_ETAT_ECHEANCES_CREANCES = {
    "VA": "Clients douteux ou litigieux",
    "VC": "Groupe et associés (créances)"
}

CODES_ETAT_ECHEANCES_DETTES = {
    "VH": "Annuité à venir",
    "VI": "Groupe et associés (dettes)"
}

# --- Dictionnaire pour l'extraction par CODES (AFFECTATION DU RÉSULTAT 2058-C-SD) ---
CODES_AFFECTATION_RESULTAT = {
    "ZE": "Dividendes"
}

CODES_RENSEIGNEMENTS_DIVERS = {
    "YQ": "Engagements de crédit-bail mobilier",
    "YR": "Engagements de crédit-bail immobilier",
    "YT": "Sous-traitance",
    "YU": "Personnel extérieur à l'entreprise"
}

# ========================================
# MOTS-CLÉS POUR EXTRACTION PAR LIBELLÉS
# ========================================

# Dictionnaire : Code → Liste de mots-clés possibles
MOTS_CLES_BILAN_ACTIF = {
    "AB": ["frais d'établissement", "frais etablissement"],
    "CX": ["frais de développement", "frais developpement"],
    "AF": ["concessions, brevets", "concessions brevets", "brevets", "licences"],
    "AH": ["fonds commercial", "fonds de commerce", "goodwill"],
    "AJ": ["autres immobilisations incorporelles", "autres immo incorporelles"],
    "AN": ["terrains"],
    "AP": ["constructions"],
    "AR": ["installations techniques", "matériel et outillage", "installations matériel outillage"],
    "AT": ["autres immobilisations corporelles", "autres immo corporelles"],
    "AV": ["immobilisations en cours", "immo en cours"],
    "BB": ["créances rattachées à des participations", "creances participations"],
    "BD": ["autres titres immobilisés", "titres immobilisés"],
    "BF": ["prêts"],
    "BH": ["autres immobilisations financières", "autres immo financières"],
    "BJ": ["total actif immobilisé", "total immo", "total immobilisations"],
    "BL": ["matières premières", "matieres premieres", "approvisionnements"],
    "BN": ["en cours de production de biens", "en-cours biens"],
    "BP": ["en cours de production de services", "en-cours services"],
    "BR": ["produits intermédiaires et finis", "produits finis"],
    "BT": ["marchandises"],
    "BX": ["clients et comptes rattachés", "creances clients", "clients"],
    "BZ": ["autres créances", "autres creances"],
    "CD": ["valeurs mobilières de placement", "vmp", "valeurs mobilieres"],
    "CF": ["disponibilités", "disponibilites", "trésorerie", "tresorerie"],
    "CH": ["charges constatées d'avance", "charges constatees avance"],
    "CJ": ["total actif circulant", "total circulant"],
    "CO": ["total général", "total general", "total bilan", "total actif"],
}

MOTS_CLES_BILAN_PASSIF = {
    "DA": ["capital social", "capital individuel"],
    "DB": ["primes d'émission", "primes emission", "primes de fusion"],
    "DD": ["réserve légale", "reserve legale"],
    "DE": ["réserves statutaires", "reserves statutaires"],
    "DF": ["réserves réglementées", "reserves reglementees"],
    "DG": ["autres réserves", "autres reserves"],
    "DH": ["report à nouveau", "report nouveau"],
    "DI": ["résultat de l'exercice", "resultat exercice", "bénéfice", "benefice", "perte"],
    "DJ": ["subventions d'investissement", "subventions investissement"],
    "DK": ["provisions réglementées", "provisions reglementees"],
    "DL": ["total capitaux propres", "total capitaux", "capitaux propres"],
    "DO": ["total autres fonds propres", "autres fonds propres"],
    "DP": ["provisions pour risques", "provisions risques"],
    "DQ": ["provisions pour charges", "provisions charges"],
    "DR": ["total provisions", "total provisions risques charges"],
    "DS": ["emprunts obligataires convertibles", "obligations convertibles"],
    "DT": ["autres emprunts obligataires", "emprunts obligataires"],
    "DU": ["emprunts et dettes auprès des établissements de crédit", "emprunts etablissements credit", "dettes bancaires"],
    "DV": ["emprunts et dettes financières divers", "dettes financieres divers"],
    "DW": ["avances et acomptes reçus", "avances recues", "acomptes recus"],
    "DX": ["dettes fournisseurs", "fournisseurs et comptes rattachés", "fournisseurs"],
    "DY": ["dettes fiscales et sociales", "dettes fiscales sociales"],
    "DZ": ["dettes sur immobilisations", "dettes immo"],
    "EA": ["autres dettes"],
    "EB": ["produits constatés d'avance", "produits constates avance"],
    "EC": ["total dettes"],
    "EE": ["total général", "total general", "total passif"],
}

MOTS_CLES_COMPTE_RESULTAT = {
    "FL": ["chiffre d'affaires nets", "chiffre affaires nets", "ca nets", "ca total", "chiffre d'affaires"],
    "FM": ["production stockée", "production stockee", "variation stocks production"],
    "FN": ["production immobilisée", "production immobilisee"],
    "FO": ["subventions d'exploitation", "subventions exploitation"],
    "FP": ["reprises sur amortissements et provisions", "reprises amortissements provisions"],
    "FQ": ["autres produits"],
    "FR": ["total des produits d'exploitation", "total produits exploitation"],
    "FS": ["achats de marchandises", "achats marchandises"],
    "FT": ["variation de stocks marchandises", "variation stocks marchandises"],
    "FU": ["achats de matières premières", "achats matieres premieres", "achats approvisionnements"],
    "FV": ["variation de stocks matières", "variation stocks matieres"],
    "FW": ["autres achats et charges externes", "autres achats charges externes", "charges externes"],
    "FX": ["impôts, taxes", "impots taxes"],
    "FY": ["salaires et traitements", "salaires"],
    "FZ": ["charges sociales"],
    "GA": ["dotations aux amortissements sur immobilisations", "dotations amortissements immo"],
    "GB": ["dotations aux provisions sur immobilisations", "dotations provisions immo"],
    "GC": ["dotations aux provisions sur actif circulant", "dotations provisions actif circulant"],
    "GD": ["dotations aux provisions pour risques et charges", "dotations provisions risques"],
    "GE": ["autres charges"],
    "GF": ["total des charges d'exploitation", "total charges exploitation"],
    "GG": ["résultat d'exploitation", "resultat exploitation"],
    "GH": ["bénéfice attribué", "benefice attribue"],
    "GI": ["perte supportée", "perte supportee"],
    "GM": ["reprises sur provisions financières", "reprises provisions financieres"],
    "GQ": ["dotations financières", "dotations financieres"],
    "GU": ["total des charges financières", "total charges financieres", "charges financières"],
    "HA": ["produits exceptionnels sur opérations de gestion", "produits exceptionnels gestion"],
    "HB": ["produits exceptionnels sur opérations en capital", "produits exceptionnels capital"],
    "HC": ["reprises sur provisions exceptionnelles", "reprises provisions exceptionnelles"],
    "HF": ["charges exceptionnelles sur opérations en capital", "charges exceptionnelles capital"],
    "HG": ["dotations exceptionnelles", "dotations exceptionnelles provisions"],
    "HI": ["résultat exceptionnel", "resultat exceptionnel"],
    "HJ": ["participation des salariés", "participation salaries"],
    "HK": ["impôts sur les bénéfices", "impots benefices", "is"],
    "HN": ["bénéfice ou perte", "resultat net", "benefice", "perte"],
    "A1": ["transfert de charges", "transferts charges"],
}

MOTS_CLES_ETAT_ECHEANCES = {
    "VA": ["clients douteux", "creances douteuses", "clients litigieux"],
    "VC": ["groupe et associés", "groupe associes", "comptes courants actif"],
    "VH": ["annuité à venir", "annuite venir", "annuites"],
    "VI": ["groupe et associés", "groupe associes", "comptes courants passif"],
}

MOTS_CLES_AFFECTATION = {
    "ZE": ["dividendes"],
    "YQ": ["engagements de crédit-bail mobilier", "credit-bail mobilier", "credit bail mobilier"],
    "YR": ["engagements de crédit-bail immobilier", "credit-bail immobilier", "credit bail immobilier"],
    "YT": ["sous-traitance"],
    "YU": ["personnel extérieur", "personnel exterieur", "intérim", "interim"],
}

# --- Dictionnaire pour l'extraction par LIBELLÉS (AU CAS OÙ) ---
LIBELLES_BILAN_ACTIF = {
    "Frais d'établissement": "Frais d'établissement", "Frais de développement": "Frais de développement",
    "Concessions, brevets et droits similaires": "Concessions, brevets et droits similaires", "Fonds commercial": "Fonds commercial",
    "Autres immobilisations incorporelles": "Autres immobilisations incorporelles", "Avances et acomptes sur immobilisations incorporelles": "Avances et acomptes sur immobilisations incorporelles",
    "Terrains": "Terrains", "Constructions": "Constructions", "Installations techniques, matériel et outillage industriels": "Installations techniques, matériel et outillage industriels",
    "Autres immobilisations corporelles": "Autres immobilisations corporelles", "Immobilisations en cours": "Immobilisations en cours", "Avances et acomptes": "Avances et acomptes",
    "Participations évaluées selon la méthode de mise en équivalence": "Participations évaluées selon la méthode de mise en équivalence", "Autres participations": "Autres participations",
    "Créances rattachées à des participations": "Créances rattachées à des participations", "Autres titres immobilisés": "Autres titres immobilisés", "Prêts": "Prêts",
    "Autres immobilisations financières": "Autres immobilisations financières", "TOTAL ACTIF IMMOBILISÉ": "TOTAL ACTIF IMMOBILISÉ", "Matières premières, approvisionnements": "Matières premières, approvisionnements",
    "En cours de production de biens": "En cours de production de biens", "En cours de production de services": "En cours de production de services",
    "Produits intermédiaires et finis": "Produits intermédiaires et finis", "Marchandises": "Marchandises", "Avances et acomptes versés sur commandes": "Avances et acomptes versés sur commandes",
    "Clients et comptes rattachés": "Clients et comptes rattachés", "Autres créances": "Autres créances", "Capital souscrit et appelé, non versé": "Capital souscrit et appelé, non versé",
    "Valeurs mobilières de placement": "Valeurs mobilières de placement", "Disponibilités": "Disponibilités", "Charges constatées d'avance": "Charges constatées d'avance",
    "TOTAL ACTIF CIRCULANT": "TOTAL ACTIF CIRCULANT", "Frais d'émission d'emprunt à étaler": "Frais d'émission d'emprunt à étaler", "TOTAL GÉNÉRAL": "TOTAL GÉNÉRAL",
    "Écarts de conversion actif": "Écarts de conversion actif"
}

LIBELLES_BILAN_PASSIF = {
    "Capital social ou individuel": "Capital social ou individuel",
    "Primes d'émission, de fusion, d'apport…": "Primes d'émission, de fusion, d'apport…",
    "Écarts de réévaluation": "Écarts de réévaluation",
    "Réserve légale": "Réserve légale",
    "Réserves statutaires ou contractuelles": "Réserves statutaires ou contractuelles",
    "Réserves réglementées": "Réserves réglementées",
    "Autres réserves": "Autres réserves",
    "Report à nouveau": "Report à nouveau",
    "RÉSULTAT DE L'EXERCICE (bénéfice ou perte)": "RÉSULTAT DE L'EXERCICE (bénéfice ou perte)",
    "Subventions d'investissement": "Subventions d'investissement",
    "Provisions réglementées": "Provisions réglementées",
    "Produit des émissions de titres participatifs": "Produit des émissions de titres participatifs",
    "Avances conditionnées": "Avances conditionnées",
    "Provisions pour risques": "Provisions pour risques",
    "Provisions pour charges": "Provisions pour charges",
    "Emprunts obligataires convertibles": "Emprunts obligatoires convertibles",
    "Autres emprunts obligatoires": "Autres emprunts obligataires",
    "Emprunts et dettes auprès des établissements de crédit": "Emprunts et dettes auprès des établissements de crédit",
    "Emprunts et dettes financières divers": "Emprunts et dettes financières divers",
    "Avances et acomptes reçus sur commandes en cours": "Avances et acomptes reçus sur commandes en cours",
    "Dettes fournisseurs et comptes rattachés": "Dettes fournisseurs et comptes rattachés",
    "Dettes fiscales et sociales": "Dettes fiscales et sociales",
    "Dettes sur immobilisations et comptes rattachés": "Dettes sur immobilisations et comptes rattachés",
    "Autres dettes": "Autres dettes",
    "Produits constatés d'avance": "Produits constatés d'avance",
    "TOTAL (I) - Capitaux propres": "TOTAL (I) - Capitaux propres",
    "TOTAL GENERAL (I à V)": "TOTAL GENERAL (I à V)"
}

# --- Seuils pour décider si l'extraction par codes est un succès ---
SEUIL_REUSSITE_CODES = 5
SEUIL_REUSSITE_CODES_PASSIF = 5
SEUIL_REUSSITE_CODES_COMPTE_RESULTAT = 10  # Plus de codes dans le Compte de Résultat
SEUIL_REUSSITE_CODES_ETAT_ECHEANCES = 3  # Au moins 3 valeurs sur 4 pour l'État des échéances
SEUIL_REUSSITE_CODES_AFFECTATION_RESULTAT = 4  # Au moins 4 valeurs sur 5 (1 affectation + 4 renseignements)


# ============================================
# FONCTIONS OUTILS
# ============================================

def nettoyer_montant(texte):
    """Nettoie et convertit un montant textuel en nombre.
    
    Gère les montants négatifs sous plusieurs formats :
    - (1 000 000) → -1000000
    - -1 000 000 → -1000000
    - 1 000 000- → -1000000
    """
    if not texte: 
        return None
    
    texte = str(texte).strip()
    
    # Cas 1 : Montant entre parenthèses → négatif
    if texte.startswith('(') and texte.endswith(')'):
        texte = '-' + texte[1:-1]  # Enlever les parenthèses et ajouter le signe -
    
    # Cas 2 : Signe - à la fin → déplacer au début
    elif texte.endswith('-'):
        texte = '-' + texte[:-1]  # Enlever le - à la fin et le mettre au début
    
    # Nettoyer : enlever espaces, remplacer virgule par point
    texte = texte.replace(',', '.').replace(' ', '')
    
    try:
        return float(texte)
    except:
        return None


def matcher_par_mots_cles(texte, mots_cles):
    """Vérifie si un des mots-clés est présent dans le texte.
    
    Args:
        texte: Le texte à analyser
        mots_cles: Liste de mots-clés à chercher
        
    Returns:
        bool: True si au moins un mot-clé est trouvé
    """
    if not texte or not mots_cles:
        return False
    
    texte_clean = texte.lower().strip()
    
    for mot_cle in mots_cles:
        if mot_cle.lower() in texte_clean:
            return True
    
    return False


def calculer_similarite(texte1, texte2):
    """Calcule la similarité entre deux textes (fuzzy matching).
    
    Args:
        texte1: Premier texte
        texte2: Second texte
        
    Returns:
        float: Score de similarité entre 0 et 1
    """
    from difflib import SequenceMatcher
    
    if not texte1 or not texte2:
        return 0.0
    
    texte1_clean = texte1.lower().strip()
    texte2_clean = texte2.lower().strip()
    
    return SequenceMatcher(None, texte1_clean, texte2_clean).ratio()


def extraire_valeur_hybride(table, code, mots_cles, libelle_reference, index_montant, seuil_fuzzy=0.75):
    """Extrait une valeur en utilisant 3 niveaux : code, mots-clés, fuzzy.
    
    Args:
        table: Tableau extrait du PDF
        code: Code officiel à chercher (ex: "FL")
        mots_cles: Liste de mots-clés alternatifs
        libelle_reference: Libellé de référence pour fuzzy matching
        index_montant: Index de la colonne contenant le montant
        seuil_fuzzy: Seuil minimum de similarité pour fuzzy (0.0 à 1.0)
        
    Returns:
        tuple: (montant, methode_utilisee) où methode = "code" | "mots_cles" | "fuzzy" | "non_trouve"
    """
    
    for row_idx, row in enumerate(table):
        for col_idx, cell in enumerate(row):
            if not cell:
                continue
            
            cell_text = str(cell).strip()
            
            # NIVEAU 1 : Chercher le CODE exact
            if cell_text.upper() == code:
                montant_cell = row[index_montant] if len(row) > index_montant else None
                montant = nettoyer_montant(montant_cell)
                if montant is not None:
                    return montant, "code"
            
            # NIVEAU 2 : Chercher par MOTS-CLÉS
            if matcher_par_mots_cles(cell_text, mots_cles):
                montant_cell = row[index_montant] if len(row) > index_montant else None
                montant = nettoyer_montant(montant_cell)
                if montant is not None:
                    return montant, "mots_cles"
            
            # NIVEAU 3 : Chercher par FUZZY matching
            similarite = calculer_similarite(cell_text, libelle_reference)
            if similarite >= seuil_fuzzy:
                montant_cell = row[index_montant] if len(row) > index_montant else None
                montant = nettoyer_montant(montant_cell)
                if montant is not None:
                    return montant, "fuzzy"
    
    return 0, "non_trouve"


def nettoyer_montant(texte):
    """Nettoie et convertit un montant textuel en nombre.
    
    Gère les montants négatifs sous plusieurs formats :
    - (1 000 000) → -1000000
    - -1 000 000 → -1000000
    - 1 000 000- → -1000000
    """
    if not texte: 
        return None
    
    texte = str(texte).strip()
    
    # Cas 1 : Montant entre parenthèses → négatif
    if texte.startswith('(') and texte.endswith(')'):
        texte = '-' + texte[1:-1]  # Enlever les parenthèses et ajouter le signe -
    
    # Cas 2 : Signe - à la fin → déplacer au début
    elif texte.endswith('-'):
        texte = '-' + texte[:-1]  # Enlever le - à la fin et le mettre au début
    
    # Nettoyer : enlever espaces, remplacer virgule par point
    texte = texte.replace(',', '.').replace(' ', '')
    
    # Ne garder que les chiffres, le point et le signe -
    texte_clean = "".join(char for char in texte if char.isdigit() or char in ['.', '-'])
    
    if not texte_clean or texte_clean == '-': 
        return None
    
    try: 
        return float(texte_clean)
    except ValueError: 
        return None

def normaliser_texte(texte):
    """Nettoie un texte pour le rendre comparable (minuscule, sans accents, sans espaces)."""
    if not texte: return ""
    texte = str(texte).lower()
    accents = {'á': 'a', 'à': 'a', 'â': 'a', 'ä': 'a', 'ã': 'a', 'å': 'a', 'ç': 'c', 'é': 'e', 'è': 'e', 'ê': 'e', 'ë': 'e', 'í': 'i', 'ì': 'i', 'î': 'i', 'ï': 'i', 'ñ': 'n', 'ó': 'o', 'ò': 'o', 'ô': 'o', 'ö': 'o', 'õ': 'o', 'ú': 'u', 'ù': 'u', 'û': 'u', 'ü': 'u', 'ý': 'y', 'ÿ': 'y'}
    for acc_char, char in accents.items():
        texte = texte.replace(acc_char, char)
    return "".join(char for char in texte if char.isalnum())


# ============================================
# FONCTIONS D'EXTRACTION - ACTIF
# ============================================

def _trouver_colonne_net(table):
    """Trouve l'index de la colonne 'Net' dans le tableau de l'actif."""
    for row in table[:5]:
        for idx, cell in enumerate(row):
            if cell and "Net" in str(cell):
                return idx
    return None

def extraire_bilan_actif_par_codes(chemin_pdf, table_actif):
    """Extrait le Bilan Actif en cherchant les CODES dans le tableau."""
    print("   → Tentative d'extraction par CODES...")
    
    idx_net = _trouver_colonne_net(table_actif)
    if idx_net is None:
        print("   ⚠️ Colonne 'Net' introuvable.")
        return [], 0

    codes_trouves = {}
    for row in table_actif:
        if not row: continue
        for idx, cell in enumerate(row):
            if cell:
                cell_text = str(cell).strip().upper()
                if cell_text in CODES_BILAN_ACTIF:
                    montant_brut = nettoyer_montant(row[idx_net]) if idx_net < len(row) else None
                    montant = montant_brut if montant_brut is not None else 0.0
                    codes_trouves[cell_text] = montant

    nb_trouves = len([v for v in codes_trouves.values() if v != 0.0])
    print(f"   ℹ️ Codes détectés : {len(codes_trouves)} | Valeurs non-nulles : {nb_trouves}")

    resultats = [(CODES_BILAN_ACTIF[code], codes_trouves.get(code, 0)) for code in CODES_BILAN_ACTIF.keys()]
    return resultats, nb_trouves

def extraire_bilan_actif_par_libelles(chemin_pdf, table_actif):
    """Extrait le Bilan Actif en cherchant les LIBELLÉS dans le tableau (méthode de secours)."""
    print("   → Extraction par LIBELLÉS...")
    
    idx_net = _trouver_colonne_net(table_actif)
    if idx_net is None:
        print("   ⚠️ Colonne 'Net' introuvable.")
        return []

    libelles_normalises = {normaliser_texte(lib): lib for lib in LIBELLES_BILAN_ACTIF.keys()}
    libelles_trouves = {}

    for row in table_actif:
        if not row: continue
        
        libelle_trouve = None
        for cell in row:
            if cell:
                cell_normalise = normaliser_texte(str(cell).strip())
                if cell_normalise in libelles_normalises:
                    libelle_trouve = libelles_normalises[cell_normalise]
                    break

        if libelle_trouve:
            montant_brut = nettoyer_montant(row[idx_net]) if idx_net < len(row) else None
            montant = montant_brut if montant_brut is not None else 0.0
            libelles_trouves[libelle_trouve] = montant

    resultats = [(libelle, libelles_trouves.get(libelle, 0)) for libelle in LIBELLES_BILAN_ACTIF.keys()]
    return resultats


# ============================================
# FONCTIONS D'EXTRACTION - PASSIF
# ============================================

def _trouver_colonne_passif_n(table_passif):
    """Trouve l'index de la colonne 'Exercice N' dans le tableau du passif.
    Les montants sont décalés de +1 par rapport à l'en-tête 'Exercice N'."""
    print("🔍 Recherche de la colonne 'Exercice N' pour le Passif...")
    for row in table_passif:
        if any("Exercice N" in str(cell) for cell in row):
            for i, cell in enumerate(row):
                if cell and "Exercice N" in str(cell):
                    idx_montant = i + 1  # ← DÉCALAGE DE +1 !
                    print(f"   ✓ En-tête 'Exercice N' trouvé à l'index {i}.")
                    print(f"   ✓ La colonne des montants sera donc l'index : {idx_montant}\n")
                    return idx_montant
    print("❌ En-tête 'Exercice N' non trouvé.")
    return None

def extraire_bilan_passif_par_codes(chemin_pdf, table_passif):
    """Extrait le Bilan Passif en cherchant les CODES dans le tableau."""
    print("   → Tentative d'extraction par CODES...")
    
    idx_passif_n = _trouver_colonne_passif_n(table_passif)
    if idx_passif_n is None:
        print("   ⚠️ Colonne 'Exercice N' introuvable.")
        return [], 0

    codes_trouves = {}
    for row in table_passif:
        if not row: continue
        for idx, cell in enumerate(row):
            if cell:
                cell_text = str(cell).strip().upper()
                if cell_text in CODES_BILAN_PASSIF:
                    montant_brut = nettoyer_montant(row[idx_passif_n]) if idx_passif_n < len(row) else None
                    montant = montant_brut if montant_brut is not None else 0.0
                    codes_trouves[cell_text] = montant

    nb_trouves = len([v for v in codes_trouves.values() if v != 0.0])
    print(f"   ℹ️ Codes détectés : {len(codes_trouves)} | Valeurs non-nulles : {nb_trouves}")

    resultats = [(CODES_BILAN_PASSIF[code], codes_trouves.get(code, 0)) for code in CODES_BILAN_PASSIF.keys()]
    return resultats, nb_trouves

def extraire_bilan_passif_par_libelles(chemin_pdf, table_passif):
    """Extrait le Bilan Passif en cherchant les LIBELLÉS dans le tableau (méthode de secours)."""
    print("   → Extraction par LIBELLÉS...")
    
    idx_passif_n = _trouver_colonne_passif_n(table_passif)
    if idx_passif_n is None:
        print("   ⚠️ Colonne 'Exercice N' introuvable.")
        return []

    libelles_normalises = {normaliser_texte(lib): lib for lib in LIBELLES_BILAN_PASSIF.keys()}
    libelles_trouves = {}

    for row in table_passif:
        if not row: continue
        
        libelle_trouve = None
        for cell in row:
            if cell:
                cell_normalise = normaliser_texte(str(cell).strip())
                if cell_normalise in libelles_normalises:
                    libelle_trouve = libelles_normalises[cell_normalise]
                    break

        if libelle_trouve:
            montant_brut = nettoyer_montant(row[idx_passif_n]) if idx_passif_n < len(row) else None
            montant = montant_brut if montant_brut is not None else 0.0
            libelles_trouves[libelle_trouve] = montant

    resultats = [(libelle, libelles_trouves.get(libelle, 0)) for libelle in LIBELLES_BILAN_PASSIF.keys()]
    return resultats


# ============================================
# FONCTIONS D'EXTRACTION - COMPTE DE RÉSULTAT
# ============================================

def _trouver_colonne_compte_resultat_page1(table_cr):
    """Trouve l'index de la colonne avec les montants dans la PAGE 1 du Compte de Résultat.
    
    Logique GPS pour PAGE 3 du PDF :
    1. Cherche "Exercice N" (devrait être à l'index 3)
    2. Cherche "TOTAL" (devrait être à l'index 7)
    3. Les montants sont à index 8 (décalage de +1 par rapport à TOTAL)
    """
    
    # Étape 1 : Chercher "Exercice N" dans les premières lignes
    idx_exercice_n = None
    for row in table_cr[:5]:
        for idx, cell in enumerate(row):
            if cell and "Exercice N" in str(cell).strip():
                idx_exercice_n = idx
                print(f"   ℹ️ [PAGE 1] 'Exercice N' trouvé à l'index {idx}")
                break
        if idx_exercice_n is not None:
            break
    
    # Étape 2 : Chercher "TOTAL" dans les premières lignes (accepte TOTAL, Total, total)
    idx_total = None
    for row in table_cr[:5]:
        for idx, cell in enumerate(row):
            if cell:
                cell_text = str(cell).strip().upper()  # ← Convertir en majuscules
                if "TOTAL" in cell_text:  # ← Utiliser "in" au lieu de "=="
                    idx_total = idx
                    print(f"   ℹ️ [PAGE 1] 'TOTAL' trouvé à l'index {idx} (valeur: '{str(cell).strip()}')")
                    break
        if idx_total is not None:
            break
    
    # Étape 3 : Déduire l'index des montants
    if idx_total is not None:
        # Les montants sont décalés de +1 par rapport à TOTAL
        idx_montants = idx_total + 1
        print(f"   ✓ [PAGE 1] Colonne des montants déduite : index {idx_montants}")
        return idx_montants
    elif idx_exercice_n is not None:
        # Si on a trouvé "Exercice N", on peut essayer de déduire
        # (mais c'est moins fiable)
        print(f"   ⚠️ [PAGE 1] 'TOTAL' non trouvé, utilisation de 'Exercice N' comme référence")
        return idx_exercice_n
    else:
        print("   ⚠️ [PAGE 1] Impossible de trouver 'Exercice N' ou 'TOTAL'")
        return None

def _trouver_colonne_compte_resultat_page2(table_cr):
    """Trouve l'index de la colonne avec les montants dans la PAGE 2 du Compte de Résultat.
    
    Logique GPS pour PAGE 4 du PDF :
    1. Cherche "Exercice N" (devrait être à l'index 12)
    2. Les montants sont à index 13 (décalage de +1 par rapport à Exercice N)
    """
    
    # Chercher "Exercice N" dans les premières lignes
    idx_exercice_n = None
    for row in table_cr[:10]:  # Chercher dans les 10 premières lignes
        for idx, cell in enumerate(row):
            if cell and "Exercice N" in str(cell).strip():
                idx_exercice_n = idx
                print(f"   ℹ️ [PAGE 2] 'Exercice N' trouvé à l'index {idx}")
                break
        if idx_exercice_n is not None:
            break
    
    # Déduire l'index des montants
    if idx_exercice_n is not None:
        # Les montants sont décalés de +1 par rapport à Exercice N
        idx_montants = idx_exercice_n + 1
        print(f"   ✓ [PAGE 2] Colonne des montants déduite : index {idx_montants}")
        return idx_montants
    else:
        print("   ⚠️ [PAGE 2] Impossible de trouver 'Exercice N'")
        return None

def extraire_compte_resultat_par_codes(chemin_pdf, pdf_obj):
    """Extrait le Compte de Résultat en cherchant les CODES dans les tableaux des DEUX pages.
    
    Le Compte de Résultat est sur 2 pages :
    - PAGE 3 du PDF (page 1 du CR) : Codes FA à GW
    - PAGE 4 du PDF (page 2 du CR) : Codes HA à HN + HP, HQ, A1
    """
    print("   → Tentative d'extraction par CODES (2 pages)...")
    
    codes_trouves = {}
    
    # ========================================
    # ÉTAPE 1 : TRAITER LA PAGE 1 (PAGE 3 DU PDF)
    # ========================================
    print("\n   📄 Traitement de la PAGE 1 du Compte de Résultat...")
    
    # Trouver la page 1
    cr_page1_index = -1
    for i, page in enumerate(pdf_obj.pages):
        text = page.extract_text()
        if text and ("Ventes de marchandises" in text or "Ventes" in text):
            cr_page1_index = i
            print(f"   ✓ Page 1 identifiée : page {i + 1} du PDF")
            break
    
    if cr_page1_index != -1:
        tables_page1 = pdf_obj.pages[cr_page1_index].extract_tables()
        if tables_page1:
            table_page1 = tables_page1[0]
            idx_montant_page1 = _trouver_colonne_compte_resultat_page1(table_page1)
            
            if idx_montant_page1 is not None:
                # Extraire les codes de la page 1
                for row in table_page1:
                    if not row: continue
                    for idx, cell in enumerate(row):
                        if cell:
                            cell_text = str(cell).strip().upper()
                            if cell_text in CODES_COMPTE_RESULTAT:
                                montant_brut = nettoyer_montant(row[idx_montant_page1]) if idx_montant_page1 < len(row) else None
                                montant = montant_brut if montant_brut is not None else 0.0
                                codes_trouves[cell_text] = montant
    
    # ========================================
    # ÉTAPE 2 : TRAITER LA PAGE 2 (PAGE 4 DU PDF)
    # ========================================
    print("\n   📄 Traitement de la PAGE 2 du Compte de Résultat...")
    
    # Trouver la page 2
    cr_page2_index = -1
    for i, page in enumerate(pdf_obj.pages):
        text = page.extract_text()
        if text and ("Produits exceptionnels" in text or "PRODUITS EXCEPTIONNELS" in text):
            cr_page2_index = i
            print(f"   ✓ Page 2 identifiée : page {i + 1} du PDF")
            break
    
    if cr_page2_index != -1:
        tables_page2 = pdf_obj.pages[cr_page2_index].extract_tables()
        if tables_page2:
            table_page2 = tables_page2[0]
            idx_montant_page2 = _trouver_colonne_compte_resultat_page2(table_page2)
            
            if idx_montant_page2 is not None:
                # Extraire les codes de la page 2
                for row in table_page2:
                    if not row: continue
                    for idx, cell in enumerate(row):
                        if cell:
                            cell_text = str(cell).strip().upper()
                            if cell_text in CODES_COMPTE_RESULTAT:
                                montant_brut = nettoyer_montant(row[idx_montant_page2]) if idx_montant_page2 < len(row) else None
                                montant = montant_brut if montant_brut is not None else 0.0
                                codes_trouves[cell_text] = montant

    nb_trouves = len([v for v in codes_trouves.values() if v != 0.0])
    print(f"\n   ℹ️ Codes détectés : {len(codes_trouves)} | Valeurs non-nulles : {nb_trouves}")

    resultats = [(CODES_COMPTE_RESULTAT[code], codes_trouves.get(code, 0)) for code in CODES_COMPTE_RESULTAT.keys()]
    return resultats, nb_trouves

def extraire_compte_resultat_par_libelles(chemin_pdf, table_cr):
    """Extrait le Compte de Résultat en cherchant les LIBELLÉS dans le tableau (méthode de secours)."""
    print("   → Extraction par LIBELLÉS...")
    
    idx_montant = _trouver_colonne_compte_resultat(table_cr)
    if idx_montant is None:
        print("   ⚠️ Colonne de montants introuvable.")
        return []

    libelles_normalises = {normaliser_texte(lib): lib for lib in CODES_COMPTE_RESULTAT.values()}
    libelles_trouves = {}

    for row in table_cr:
        if not row: continue
        
        libelle_trouve = None
        for cell in row:
            if cell:
                cell_normalise = normaliser_texte(str(cell).strip())
                if cell_normalise in libelles_normalises:
                    libelle_trouve = libelles_normalises[cell_normalise]
                    break

        if libelle_trouve:
            montant_brut = nettoyer_montant(row[idx_montant]) if idx_montant < len(row) else None
            montant = montant_brut if montant_brut is not None else 0.0
            libelles_trouves[libelle_trouve] = montant

    resultats = [(libelle, libelles_trouves.get(libelle, 0)) for libelle in CODES_COMPTE_RESULTAT.values()]
    return resultats


# ============================================
# FONCTIONS PRINCIPALES
# ============================================

# Dictionnaires inverses pour retrouver les codes à partir des libellés
LIBELLE_TO_CODE_ACTIF = {v: k for k, v in CODES_BILAN_ACTIF.items()}
LIBELLE_TO_CODE_PASSIF = {v: k for k, v in CODES_BILAN_PASSIF.items()}
LIBELLE_TO_CODE_CR = {v: k for k, v in CODES_COMPTE_RESULTAT.items()}
LIBELLE_TO_CODE_ECHEANCES_CREANCES = {v: k for k, v in CODES_ETAT_ECHEANCES_CREANCES.items()}
LIBELLE_TO_CODE_ECHEANCES_DETTES = {v: k for k, v in CODES_ETAT_ECHEANCES_DETTES.items()}
LIBELLE_TO_CODE_AFFECTATION = {v: k for k, v in CODES_AFFECTATION_RESULTAT.items()}
LIBELLE_TO_CODE_RENSEIGNEMENTS = {v: k for k, v in CODES_RENSEIGNEMENTS_DIVERS.items()}

def creer_fichier_excel(donnees_par_annee, nom_fichier):
    """Crée le fichier Excel avec UN SEUL onglet structuré par catégories.
    
    Args:
        donnees_par_annee: Dict avec structure {
            'annee1': {'actif': [...], 'passif': [...], 'cr': [...], 'echeances': [...], 'affectation': [...]},
            'annee2': {...}
        }
        nom_fichier: Path du fichier Excel à créer
    """
    print(f"📊 Création du fichier : {nom_fichier.name}")
    wb = openpyxl.Workbook()
    
    # Trier les années chronologiquement
    annees_triees = sorted(donnees_par_annee.keys())
    print(f"   📅 Années détectées : {', '.join(annees_triees)}")
    
    # ========================================
    # ONGLET UNIQUE avec colonne Catégorie
    # ========================================
    ws = wb.active
    ws.title = "Données Fiscales"
    
    # En-têtes
    ws['A1'] = "Catégorie"
    ws['A1'].font = openpyxl.styles.Font(bold=True, size=12)
    ws['B1'] = "Code"
    ws['B1'].font = openpyxl.styles.Font(bold=True, size=12)
    ws['C1'] = "Libellé"
    ws['C1'].font = openpyxl.styles.Font(bold=True, size=12)

    for col_idx, annee in enumerate(annees_triees, start=4):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = annee
        cell.font = openpyxl.styles.Font(bold=True, size=12)
    
    # Ligne actuelle
    current_row = 2
    
    # ========================================
    # SECTION 1: BILAN ACTIF
    # ========================================
    if annees_triees:
        premiere_annee = annees_triees[0]
        donnees_actif = donnees_par_annee[premiere_annee].get('actif', [])

        for libelle, _ in donnees_actif:
            ws[f'A{current_row}'] = "BILAN ACTIF"
            ws[f'B{current_row}'] = LIBELLE_TO_CODE_ACTIF.get(libelle, "")
            ws[f'C{current_row}'] = libelle

            # Mettre en gras les TOTAUX
            if "TOTAL" in libelle.upper():
                ws[f'A{current_row}'].font = openpyxl.styles.Font(bold=True)
                ws[f'B{current_row}'].font = openpyxl.styles.Font(bold=True)
                ws[f'C{current_row}'].font = openpyxl.styles.Font(bold=True)

            # Remplir les montants pour chaque année
            for col_idx, annee in enumerate(annees_triees, start=4):
                donnees_annee = donnees_par_annee[annee].get('actif', [])
                montant = next((m for l, m in donnees_annee if l == libelle), 0)

                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = montant
                cell.number_format = '#,##0.00'

                if "TOTAL" in libelle.upper():
                    cell.font = openpyxl.styles.Font(bold=True)

            current_row += 1
    
    # Ligne vide entre sections
    current_row += 1
    
    # ========================================
    # SECTION 2: BILAN PASSIF
    # ========================================
    if annees_triees:
        premiere_annee = annees_triees[0]
        donnees_passif = donnees_par_annee[premiere_annee].get('passif', [])

        for libelle, _ in donnees_passif:
            ws[f'A{current_row}'] = "BILAN PASSIF"
            ws[f'B{current_row}'] = LIBELLE_TO_CODE_PASSIF.get(libelle, "")
            ws[f'C{current_row}'] = libelle

            # Mettre en gras les TOTAUX
            if "TOTAL" in libelle.upper():
                ws[f'A{current_row}'].font = openpyxl.styles.Font(bold=True)
                ws[f'B{current_row}'].font = openpyxl.styles.Font(bold=True)
                ws[f'C{current_row}'].font = openpyxl.styles.Font(bold=True)

            # Remplir les montants pour chaque année
            for col_idx, annee in enumerate(annees_triees, start=4):
                donnees_annee = donnees_par_annee[annee].get('passif', [])
                montant = next((m for l, m in donnees_annee if l == libelle), 0)

                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = montant
                cell.number_format = '#,##0.00'

                if "TOTAL" in libelle.upper():
                    cell.font = openpyxl.styles.Font(bold=True)

            current_row += 1
    
    # Ligne vide entre sections
    current_row += 1
    
    # ========================================
    # SECTION 3: COMPTE DE RÉSULTAT
    # ========================================
    if annees_triees:
        premiere_annee = annees_triees[0]
        donnees_cr = donnees_par_annee[premiere_annee].get('cr', [])

        for libelle, _ in donnees_cr:
            ws[f'A{current_row}'] = "COMPTE RÉSULTAT"
            ws[f'B{current_row}'] = LIBELLE_TO_CODE_CR.get(libelle, "")
            ws[f'C{current_row}'] = libelle

            # Mettre en gras les TOTAUX et RÉSULTATS
            if any(keyword in libelle.upper() for keyword in ["TOTAL", "RÉSULTAT", "CHIFFRE D'AFFAIRES", "BÉNÉFICE", "PERTE"]):
                ws[f'A{current_row}'].font = openpyxl.styles.Font(bold=True)
                ws[f'B{current_row}'].font = openpyxl.styles.Font(bold=True)
                ws[f'C{current_row}'].font = openpyxl.styles.Font(bold=True)

            # Remplir les montants pour chaque année
            for col_idx, annee in enumerate(annees_triees, start=4):
                donnees_annee = donnees_par_annee[annee].get('cr', [])
                montant = next((m for l, m in donnees_annee if l == libelle), 0)

                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = montant
                cell.number_format = '#,##0.00'

                if any(keyword in libelle.upper() for keyword in ["TOTAL", "RÉSULTAT", "CHIFFRE D'AFFAIRES", "BÉNÉFICE", "PERTE"]):
                    cell.font = openpyxl.styles.Font(bold=True)

            current_row += 1
    
    # Ligne vide entre sections
    current_row += 1
    
    # ========================================
    # SECTION 4: ÉTAT DES ÉCHÉANCES
    # ========================================
    if annees_triees:
        premiere_annee = annees_triees[0]
        donnees_echeances = donnees_par_annee[premiere_annee].get('echeances', [])

        for libelle, _ in donnees_echeances:
            ws[f'A{current_row}'] = "ÉCHÉANCES"
            # Chercher le code dans les deux dictionnaires d'échéances
            code = LIBELLE_TO_CODE_ECHEANCES_CREANCES.get(libelle) or LIBELLE_TO_CODE_ECHEANCES_DETTES.get(libelle, "")
            ws[f'B{current_row}'] = code
            ws[f'C{current_row}'] = libelle

            # Remplir les montants pour chaque année
            for col_idx, annee in enumerate(annees_triees, start=4):
                donnees_annee = donnees_par_annee[annee].get('echeances', [])
                montant = next((m for l, m in donnees_annee if l == libelle), 0)

                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = montant
                cell.number_format = '#,##0.00'

            current_row += 1
    
    # Ligne vide entre sections
    current_row += 1
    
    # ========================================
    # SECTION 5: AFFECTATION & RENSEIGNEMENTS
    # ========================================
    if annees_triees:
        premiere_annee = annees_triees[0]
        donnees_affectation = donnees_par_annee[premiere_annee].get('affectation', [])

        for libelle, _ in donnees_affectation:
            ws[f'A{current_row}'] = "AFFECTATION"
            # Chercher le code dans les dictionnaires d'affectation et renseignements
            code = LIBELLE_TO_CODE_AFFECTATION.get(libelle) or LIBELLE_TO_CODE_RENSEIGNEMENTS.get(libelle, "")
            ws[f'B{current_row}'] = code
            ws[f'C{current_row}'] = libelle

            # Remplir les montants pour chaque année
            for col_idx, annee in enumerate(annees_triees, start=4):
                donnees_annee = donnees_par_annee[annee].get('affectation', [])
                montant = next((m for l, m in donnees_annee if l == libelle), 0)

                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = montant
                cell.number_format = '#,##0.00'

            current_row += 1
    
    # Ajuster la largeur des colonnes
    ws.column_dimensions['A'].width = 20  # Catégorie
    ws.column_dimensions['B'].width = 10  # Code
    ws.column_dimensions['C'].width = 60  # Libellé
    for col_idx in range(4, len(annees_triees) + 4):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15

    # Figer les en-têtes (ligne 1) et les colonnes Catégorie + Code + Libellé
    ws.freeze_panes = 'D2'
    
    # ========================================
    # ONGLET 2: ANALYSE FINANCIÈRE
    # ========================================
    print("   📊 Calcul des ratios financiers...")
    ratios_par_annee = calculer_ratios_financiers(donnees_par_annee)
    
    ws_analyse = wb.create_sheet("Analyse Financière")
    
    # En-têtes
    ws_analyse['A1'] = "Indicateur"
    ws_analyse['A1'].font = openpyxl.styles.Font(bold=True, size=12)
    
    for col_idx, annee in enumerate(annees_triees, start=2):
        cell = ws_analyse.cell(row=1, column=col_idx)
        cell.value = annee
        cell.font = openpyxl.styles.Font(bold=True, size=12)
    
    current_row = 2
    
    # Structure des données à afficher
    structure_analyse = [
        ("=== ACTIVITÉ & RENTABILITÉ (K€) ===", None, None),
        ("Durée (en mois)", "duree_mois", None),
        ("CA", "ca", None),
        ("Production stockée + immobilisée", "prod_stockee_immo", None),
        ("Production globale", "prod_globale", None),
        ("AACE", "aace", None),
        ("Dont sous-traitance", "sous_traitance", None),
        ("% Production globale", "pct_sous_traitance", "%"),
        ("Production interne", "prod_interne", None),
        ("Consommation de matières premières et marchandises", "conso_matieres", None),
        ("% production interne", "pct_conso_matieres", "%"),
        ("Charge de personnel", "charge_personnel", None),
        ("Intérim", "interim", None),
        ("% production interne", "pct_charge_personnel", "%"),
        ("EBE", "ebe", None),
        ("% production interne", "pct_ebe", "%"),
        ("Résultat d'exploitation", "resultat_exploitation", None),
        ("% production interne", "pct_resultat_exploitation", "%"),
        ("Charges financières", "charges_financieres", None),
        ("%EBE", "pct_charges_financieres", "%"),
        ("Résultat exceptionnel", "resultat_exceptionnel", None),
        ("Résultat net", "resultat_net", None),
        ("%PI", "pct_resultat_net", "%"),
        ("CAF (y.c crédit bail)", "caf", None),
        ("", None, None),  # Ligne vide
        ("=== BILAN (K€) ===", None, None),
        ("Durée (en mois)", "duree_mois", None),
        ("Non-valeurs", "non_valeurs", None),
        ("Total bilan", "total_bilan", None),
        ("Capitaux propres", "capitaux_propres", None),
        ("Solvabilité (%)", "solvabilite", "%"),
        ("Couverture de l'activité (%)", "couverture_activite", "%"),
        ("Dette brute", "dette_brute", None),
        ("Gearing brut", "gearing_brut", None),
        ("Leverage brut", "leverage_brut", None),
        ("Dont MLT", "dont_mlt", None),
        ("Dont crédit bail", "dont_cb", None),
        ("Capacité de remboursement", "capacite_remboursement", None),
        ("Annuités à venir", "annuites", None),
        ("Couverture des annuités à venir avec la CAF", "couverture_annuites", None),
        ("Dette nette", "dette_nette", None),
        ("Gearing net", "gearing_net", None),
        ("Leverage net", "leverage_net", None),
        ("C/C Actif", "cc_actif", None),
        ("C/C Passif", "cc_passif", None),
        ("Dividendes", "dividendes", None),
        ("", None, None),  # Ligne vide
        ("=== CYCLE D'EXPLOITATION (K€) ===", None, None),
        ("Durée (en mois)", "duree_mois", None),
        ("FRNG", "frng", None),
        ("BFR", "bfr", None),
        ("Dont BFRE", "bfre", None),
        ("Nb jours", "nb_jours_bfre", "jours"),
        ("Dont Stocks", "stocks", None),
        ("Nb jours", "nb_jours_stocks", "jours"),
        ("Dont créances clients", "creances_clients", None),
        ("Nb jours", "nb_jours_creances", "jours"),
        ("% créances douteuses", "pct_creances_douteuses", "%"),
        ("% créances douteuses provisionnées", "pct_creances_douteuses_prov", "%"),
        ("Dont dettes fournisseurs", "dettes_fournisseurs", None),
        ("Nb jours", "nb_jours_fournisseurs", "jours"),
        ("Trésorerie nette", "tresorerie_nette", None),
    ]
    
    # Remplir les données
    for libelle, cle_ratio, format_type in structure_analyse:
        ws_analyse[f'A{current_row}'] = libelle
        
        # Titres de sections en gras
        if libelle.startswith("==="):
            ws_analyse[f'A{current_row}'].font = openpyxl.styles.Font(bold=True, size=12)
            current_row += 1
            continue
        
        # Ligne vide
        if not libelle:
            current_row += 1
            continue
        
        # Remplir les valeurs pour chaque année
        for col_idx, annee in enumerate(annees_triees, start=2):
            if cle_ratio and annee in ratios_par_annee:
                valeur = ratios_par_annee[annee].get(cle_ratio, 0)
                
                cell = ws_analyse.cell(row=current_row, column=col_idx)
                cell.value = valeur
                
                # Format selon le type
                if format_type == "%":
                    cell.number_format = '0.00"%"'
                elif format_type == "jours":
                    cell.number_format = '0.0'
                else:
                    cell.number_format = '#,##0.00'
        
        current_row += 1
    
    # Ajuster largeur colonnes
    ws_analyse.column_dimensions['A'].width = 60
    for col_idx in range(2, len(annees_triees) + 2):
        ws_analyse.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
    
    # Figer en-têtes
    ws_analyse.freeze_panes = 'B2'
    
    wb.save(nom_fichier)
    print(f"✅ Fichier créé avec 2 onglets (Données + Analyse) et {len(annees_triees)} année(s)\n")


def calculer_ratios_financiers(donnees_par_annee):
    """Calcule les ratios financiers à partir des données extraites.
    
    Returns:
        dict: {annee: {ratio: valeur}}
    """
    ratios_par_annee = {}
    
    for annee, donnees in donnees_par_annee.items():
        ratios = {}
        
        # Fonction helper pour récupérer une valeur par code
        def get_valeur(section, code):
            """Récupère la valeur d'un code dans une section."""
            if section not in donnees:
                return 0
            for libelle, montant in donnees[section]:
                # Chercher dans les dictionnaires de codes
                if section == 'actif' and code in CODES_BILAN_ACTIF:
                    if libelle == CODES_BILAN_ACTIF[code]:
                        return montant or 0
                elif section == 'passif' and code in CODES_BILAN_PASSIF:
                    if libelle == CODES_BILAN_PASSIF[code]:
                        return montant or 0
                elif section == 'cr' and code in CODES_COMPTE_RESULTAT:
                    if libelle == CODES_COMPTE_RESULTAT[code]:
                        return montant or 0
                elif section == 'echeances':
                    if code == 'VA' and libelle == "Clients douteux ou litigieux":
                        return montant or 0
                    elif code == 'VC' and libelle == "Groupe et associés (créances)":
                        return montant or 0
                    elif code == 'VH' and libelle == "Annuité à venir":
                        return montant or 0
                    elif code == 'VI' and libelle == "Groupe et associés (dettes)":
                        return montant or 0
                elif section == 'affectation':
                    if code == 'ZE' and libelle == "Dividendes":
                        return montant or 0
                    elif code == 'YQ' and libelle == "Engagements de crédit-bail mobilier":
                        return montant or 0
                    elif code == 'YR' and libelle == "Engagements de crédit-bail immobilier":
                        return montant or 0
                    elif code == 'YT' and libelle == "Sous-traitance":
                        return montant or 0
                    elif code == 'YU' and libelle == "Personnel extérieur à l'entreprise":
                        return montant or 0
            return 0
        
        # ========================================
        # SECTION 1: ACTIVITÉ & RENTABILITÉ
        # ========================================
        
        # Durée (toujours 12 mois pour l'instant)
        ratios['duree_mois'] = 12
        
        # CA
        ca = get_valeur('cr', 'FL')
        ratios['ca'] = ca
        
        # Production stockée + immobilisée
        prod_stockee = get_valeur('cr', 'FM')
        prod_immobilisee = get_valeur('cr', 'FN')
        prod_stockee_immo = prod_stockee + prod_immobilisee
        ratios['prod_stockee_immo'] = prod_stockee_immo
        
        # Production globale
        prod_globale = ca + prod_stockee_immo
        ratios['prod_globale'] = prod_globale
        
        # AACE (Autres achats et charges externes)
        aace = get_valeur('cr', 'FW')
        ratios['aace'] = aace
        
        # Sous-traitance
        sous_traitance = get_valeur('affectation', 'YT')
        ratios['sous_traitance'] = sous_traitance
        
        # % Sous-traitance / Production globale
        ratios['pct_sous_traitance'] = (sous_traitance / prod_globale * 100) if prod_globale else 0
        
        # Production interne
        prod_interne = prod_globale - sous_traitance
        ratios['prod_interne'] = prod_interne
        
        # Consommation de matières premières et marchandises
        achats_matieres = get_valeur('cr', 'FU')
        var_stocks_matieres = get_valeur('cr', 'FV')
        conso_matieres = achats_matieres + var_stocks_matieres
        ratios['conso_matieres'] = conso_matieres
        
        # % Conso matières / Production interne
        ratios['pct_conso_matieres'] = (conso_matieres / prod_interne * 100) if prod_interne else 0
        
        # Charge de personnel
        salaires = get_valeur('cr', 'FY')
        charges_sociales = get_valeur('cr', 'FZ')
        charge_personnel = salaires + charges_sociales
        ratios['charge_personnel'] = charge_personnel
        
        # Intérim
        interim = get_valeur('affectation', 'YU')
        ratios['interim'] = interim
        
        # % (Charge personnel + Intérim) / Production interne
        ratios['pct_charge_personnel'] = ((charge_personnel + interim) / prod_interne * 100) if prod_interne else 0
        
        # EBE
        resultat_exploitation = get_valeur('cr', 'GG')
        benefice_attribue = get_valeur('cr', 'GH')
        perte_supportee = get_valeur('cr', 'GI')
        dot_amort = get_valeur('cr', 'GA')
        dot_prov_immo = get_valeur('cr', 'GB')
        dot_prov_actif_circ = get_valeur('cr', 'GC')
        dot_prov_risques = get_valeur('cr', 'GD')
        dotations_exploitation = dot_amort + dot_prov_immo + dot_prov_actif_circ + dot_prov_risques
        reprises = get_valeur('cr', 'FP')
        transferts_charges = get_valeur('cr', 'A1')
        
        ebe = resultat_exploitation + benefice_attribue - perte_supportee + dotations_exploitation - reprises - transferts_charges
        ratios['ebe'] = ebe
        
        # % EBE / Production interne
        ratios['pct_ebe'] = (ebe / prod_interne * 100) if prod_interne else 0
        
        # Résultat d'exploitation
        ratios['resultat_exploitation'] = resultat_exploitation
        
        # % Résultat exploitation / Production interne
        ratios['pct_resultat_exploitation'] = (resultat_exploitation / prod_interne * 100) if prod_interne else 0
        
        # Charges financières
        charges_financieres = get_valeur('cr', 'GU')
        ratios['charges_financieres'] = charges_financieres
        
        # % Charges financières / EBE
        ratios['pct_charges_financieres'] = (charges_financieres / ebe * 100) if ebe else 0
        
        # Résultat exceptionnel
        resultat_exceptionnel = get_valeur('cr', 'HI')
        ratios['resultat_exceptionnel'] = resultat_exceptionnel
        
        # Résultat net
        resultat_net = get_valeur('cr', 'HN')
        ratios['resultat_net'] = resultat_net
        
        # % Résultat net / Production interne
        ratios['pct_resultat_net'] = (resultat_net / prod_interne * 100) if prod_interne else 0
        
        # CAF (y.c crédit-bail)
        dot_financieres = get_valeur('cr', 'GQ')
        reprises_financieres = get_valeur('cr', 'GM')
        dot_exceptionnelles = get_valeur('cr', 'HG')
        reprises_exceptionnelles = get_valeur('cr', 'HC')
        charges_except_capital = get_valeur('cr', 'HF')
        produits_except_capital = get_valeur('cr', 'HB')
        cb_mobilier = get_valeur('affectation', 'YQ')
        cb_immobilier = get_valeur('affectation', 'YR')
        
        caf = (resultat_net + dotations_exploitation - reprises + transferts_charges + 
               dot_financieres - reprises_financieres + dot_exceptionnelles - reprises_exceptionnelles +
               charges_except_capital - produits_except_capital + (0.8 * cb_mobilier) + (0.6 * cb_immobilier))
        ratios['caf'] = caf
        
        # ========================================
        # SECTION 2: BILAN
        # ========================================
        
        # Non-valeurs (Total actif incorporel)
        non_valeurs = get_valeur('actif', 'BJ')  # TOTAL ACTIF IMMOBILISÉ
        ratios['non_valeurs'] = non_valeurs
        
        # Total bilan
        total_bilan = get_valeur('actif', 'CO')
        ratios['total_bilan'] = total_bilan
        
        # Capitaux propres
        capitaux_propres = get_valeur('passif', 'DL')
        ratios['capitaux_propres'] = capitaux_propres
        
        # Solvabilité (%)
        ratios['solvabilite'] = (capitaux_propres / total_bilan * 100) if total_bilan else 0
        
        # Couverture de l'activité (%)
        ratios['couverture_activite'] = (capitaux_propres / ca * 100) if ca else 0
        
        # Dette brute
        emprunts_obl_convert = get_valeur('passif', 'DS')
        autres_emprunts_obl = get_valeur('passif', 'DT')
        dette_mlt = get_valeur('passif', 'DU')
        concours_bancaires = get_valeur('passif', 'DU')  # Inclus dans DU normalement
        
        dette_brute = emprunts_obl_convert + autres_emprunts_obl + dette_mlt + cb_mobilier + cb_immobilier + concours_bancaires
        ratios['dette_brute'] = dette_brute
        
        # Gearing brut
        ratios['gearing_brut'] = (dette_brute / capitaux_propres) if capitaux_propres else 0
        
        # Leverage brut
        ratios['leverage_brut'] = (dette_brute / ebe) if ebe else 0
        
        # Dont MLT
        ratios['dont_mlt'] = dette_mlt
        
        # Dont crédit-bail
        ratios['dont_cb'] = cb_mobilier + cb_immobilier
        
        # Capacité de remboursement
        ratios['capacite_remboursement'] = (dette_brute / caf) if caf else 0
        
        # Annuités à venir
        annuites = get_valeur('echeances', 'VH')
        ratios['annuites'] = annuites
        
        # Couverture des annuités avec CAF
        ratios['couverture_annuites'] = (caf / annuites) if annuites else 0
        
        # Dette nette
        disponibilites = get_valeur('actif', 'CF')
        vmp = get_valeur('actif', 'CD')
        dette_nette = dette_brute - disponibilites - vmp
        ratios['dette_nette'] = dette_nette
        
        # Gearing net
        ratios['gearing_net'] = (dette_nette / capitaux_propres) if capitaux_propres else 0
        
        # Leverage net
        ratios['leverage_net'] = (dette_nette / ebe) if ebe else 0
        
        # C/C Actif
        cc_actif = get_valeur('echeances', 'VC')
        ratios['cc_actif'] = cc_actif
        
        # C/C Passif
        cc_passif = get_valeur('echeances', 'VI')
        ratios['cc_passif'] = cc_passif
        
        # Dividendes
        dividendes = get_valeur('affectation', 'ZE')
        ratios['dividendes'] = dividendes
        
        # ========================================
        # SECTION 3: CYCLE D'EXPLOITATION
        # ========================================
        
        # Actif immobilisé net
        actif_immo_net = get_valeur('actif', 'BJ')
        
        # Autres fonds propres
        autres_fonds_propres = get_valeur('passif', 'DO')
        
        # Provisions pour risques et charges
        provisions = get_valeur('passif', 'DR')
        
        # FRNG
        frng = (capitaux_propres + autres_fonds_propres + provisions + 
                emprunts_obl_convert + autres_emprunts_obl + dette_mlt - actif_immo_net)
        ratios['frng'] = frng
        
        # Stocks
        stocks_mp = get_valeur('actif', 'BL')
        stocks_prod_biens = get_valeur('actif', 'BN')
        stocks_prod_services = get_valeur('actif', 'BP')
        stocks_produits_finis = get_valeur('actif', 'BR')
        stocks_marchandises = get_valeur('actif', 'BT')
        stocks = stocks_mp + stocks_prod_biens + stocks_prod_services + stocks_produits_finis + stocks_marchandises
        ratios['stocks'] = stocks
        
        # Créances
        creances_clients = get_valeur('actif', 'BX')
        autres_creances = get_valeur('actif', 'BZ')
        creances = creances_clients + autres_creances
        ratios['creances'] = creances
        
        # Charges constatées d'avance
        charges_constatees_avance = get_valeur('actif', 'CH')
        
        # Dettes
        dettes_fin_divers = get_valeur('passif', 'DV')
        avances_recues = get_valeur('passif', 'DW')
        dettes_fournisseurs = get_valeur('passif', 'DX')
        dettes_fiscales_sociales = get_valeur('passif', 'DY')
        dettes_immo = get_valeur('passif', 'DZ')
        autres_dettes = get_valeur('passif', 'EA')
        produits_constates_avance = get_valeur('passif', 'EB')
        
        # BFR
        bfr = (stocks + creances - cc_actif + charges_constatees_avance - 
               dettes_fin_divers - avances_recues - dettes_fournisseurs - 
               dettes_fiscales_sociales - dettes_immo - autres_dettes - 
               cc_passif - produits_constates_avance)
        ratios['bfr'] = bfr
        
        # BFRE
        bfre = bfr - (cc_actif - cc_passif)
        ratios['bfre'] = bfre
        
        # Nb jours BFRE
        ratios['nb_jours_bfre'] = ((bfre / (1.2 * ca)) * 360) if ca else 0
        
        # Nb jours stocks
        ratios['nb_jours_stocks'] = ((stocks / (1.2 * ca)) * 360) if ca else 0
        
        # Créances clients
        ratios['creances_clients'] = creances_clients
        
        # Nb jours créances clients
        en_cours_prod = stocks_prod_biens + stocks_prod_services
        ratios['nb_jours_creances'] = ((en_cours_prod + creances_clients - avances_recues - produits_constates_avance) / 
                                       ((ca + prod_stockee) * 1.2) * 360) if (ca + prod_stockee) else 0
        
        # % Créances douteuses
        creances_douteuses = get_valeur('echeances', 'VA')
        ratios['creances_douteuses'] = creances_douteuses
        ratios['pct_creances_douteuses'] = (creances_douteuses / creances_clients * 100) if creances_clients else 0
        
        # % Créances douteuses provisionnées (on suppose que c'est calculé ailleurs)
        ratios['pct_creances_douteuses_prov'] = 0  # À compléter si provision disponible
        
        # Dettes fournisseurs
        ratios['dettes_fournisseurs'] = dettes_fournisseurs
        
        # Nb jours dettes fournisseurs
        if sous_traitance > 0:
            ratios['nb_jours_fournisseurs'] = ((dettes_fournisseurs / ((conso_matieres + sous_traitance) * 1.2)) * 360) if (conso_matieres + sous_traitance) else 0
        else:
            ratios['nb_jours_fournisseurs'] = ((dettes_fournisseurs / ((conso_matieres + aace) * 1.2)) * 360) if (conso_matieres + aace) else 0
        
        # Trésorerie nette
        tresorerie_nette = frng - bfr
        ratios['tresorerie_nette'] = tresorerie_nette
        
        ratios_par_annee[annee] = ratios
    
    return ratios_par_annee


def extraire_etat_echeances_par_codes(chemin_pdf, pdf):
    """
    Extrait l'État des échéances (2057-SD) en utilisant les codes officiels.
    
    Structure:
    - CRÉANCES: codes VA, VC à l'index 13
    - DETTES: code VC à l'index 14, code VI à l'index 10
    
    Returns:
        (list, int): Liste de tuples (libellé, montant) et nombre de valeurs trouvées
    """
    print("📊 Extraction par CODES de l'État des échéances...")
    
    donnees = []
    nb_trouves = 0
    
    # Trouver la page contenant l'État des échéances
    page_index = -1
    for i, page in enumerate(pdf.pages):
        text = page.extract_text()
        if text and ("ÉTAT DES ÉCHÉANCES" in text.upper() or "ETAT DES ECHEANCES" in text.upper()):
            page_index = i
            print(f"   ✓ Page 'État des échéances' trouvée : page {i + 1}")
            break
    
    if page_index == -1:
        print("   ❌ Page 'État des échéances' non trouvée.")
        return donnees, 0
    
    # Extraire le tableau
    tables = pdf.pages[page_index].extract_tables()
    if not tables:
        print("   ❌ Aucun tableau extrait.")
        return donnees, 0
    
    table = tables[0]
    print(f"   ✓ Tableau extrait ({len(table)} lignes, {len(table[0]) if table else 0} colonnes)")
    
    # Indicateur pour savoir si on est dans la section DETTES
    dans_section_dettes = False
    
    # Dictionnaires pour éviter les doublons
    codes_trouves_creances = set()
    codes_trouves_dettes = set()
    
    # Parcourir toutes les lignes pour trouver les codes
    for row_idx, row in enumerate(table):
        # Détecter le changement de section (plusieurs variantes)
        row_text = " ".join(str(cell) for cell in row if cell).upper()
        
        # Variantes possibles du titre DETTES
        if any(marker in row_text for marker in [
            "ÉTAT DES DETTES",
            "ETAT DES DETTES",
            "DETTES (4)",
            "ÉTAT DES DETTES",
            "B – PLUS-VALUES"  # Parfois le titre est différent
        ]):
            dans_section_dettes = True
            print(f"   ℹ️  Section DETTES détectée à la ligne {row_idx + 1}: {row_text[:80]}")
        
        # Chercher les codes dans la ligne
        for col_idx, cell in enumerate(row):
            if not cell:
                continue
            
            code = str(cell).strip().upper()
            
            # SECTION CRÉANCES
            if not dans_section_dettes and code in CODES_ETAT_ECHEANCES_CREANCES:
                # Éviter les doublons
                if code in codes_trouves_creances:
                    continue
                codes_trouves_creances.add(code)
                
                libelle = CODES_ETAT_ECHEANCES_CREANCES[code]
                montant_cell = row[13] if len(row) > 13 else None  # Index 13 pour créances
                montant = nettoyer_montant(montant_cell)
                
                if montant is not None:
                    donnees.append((libelle, montant))
                    nb_trouves += 1
                    print(f"   ✓ CRÉANCES - {code} ({libelle}) → {montant} [index 13]")
                else:
                    donnees.append((libelle, 0))
                    print(f"   ⚠️  CRÉANCES - {code} ({libelle}) → montant non trouvé [index 13]")
            
            # SECTION DETTES
            elif dans_section_dettes and code in CODES_ETAT_ECHEANCES_DETTES:
                # Éviter les doublons
                if code in codes_trouves_dettes:
                    continue
                codes_trouves_dettes.add(code)
                
                libelle = CODES_ETAT_ECHEANCES_DETTES[code]
                
                # Code VH (Annuité à venir) → index 14
                if code == "VH":
                    montant_cell = row[14] if len(row) > 14 else None
                    montant = nettoyer_montant(montant_cell)
                    
                    if montant is not None:
                        donnees.append((libelle, montant))
                        nb_trouves += 1
                        print(f"   ✓ DETTES - {code} ({libelle}) → {montant} [index 14]")
                    else:
                        donnees.append((libelle, 0))
                        print(f"   ⚠️  DETTES - {code} ({libelle}) → montant non trouvé [index 14]")
                
                # Code VI (Groupe et associés dettes) → index 10
                elif code == "VI":
                    montant_cell = row[10] if len(row) > 10 else None
                    montant = nettoyer_montant(montant_cell)
                    
                    if montant is not None:
                        donnees.append((libelle, montant))
                        nb_trouves += 1
                        print(f"   ✓ DETTES - {code} ({libelle}) → {montant} [index 10]")
                    else:
                        donnees.append((libelle, 0))
                        print(f"   ⚠️  DETTES - {code} ({libelle}) → montant non trouvé [index 10]")
    
    # Debug : afficher ce qui a été trouvé
    print(f"\n   🔍 Debug - Codes CRÉANCES trouvés : {codes_trouves_creances}")
    print(f"   🔍 Debug - Codes DETTES trouvés : {codes_trouves_dettes}")
    print(f"   🔍 Debug - Section DETTES détectée : {dans_section_dettes}")
    
    print(f"\n   📊 Total : {nb_trouves} valeur(s) trouvée(s) sur {len(CODES_ETAT_ECHEANCES_CREANCES) + len(CODES_ETAT_ECHEANCES_DETTES)}")
    return donnees, nb_trouves


def extraire_affectation_resultat_par_codes(chemin_pdf, pdf):
    """
    Extrait l'Affectation du résultat et Renseignements divers (2058-C-SD) en utilisant les codes officiels.
    
    Structure:
    - AFFECTATION: code ZE à l'index 26
    - RENSEIGNEMENTS DIVERS: codes YQ, YR, YT, YU à l'index 18
    
    Returns:
        (list, int): Liste de tuples (libellé, montant) et nombre de valeurs trouvées
    """
    print("📊 Extraction par CODES de l'Affectation du résultat et Renseignements divers...")
    
    donnees = []
    nb_trouves = 0
    
    # Trouver la page contenant l'Affectation du résultat
    page_index = -1
    for i, page in enumerate(pdf.pages):
        text = page.extract_text()
        if text and ("AFFECTATION DU RÉSULTAT" in text.upper() or 
                    "AFFECTATION DU RESULTAT" in text.upper() or
                    "RENSEIGNEMENTS DIVERS" in text.upper()):
            page_index = i
            print(f"   ✓ Page 'Affectation du résultat' trouvée : page {i + 1}")
            break
    
    if page_index == -1:
        print("   ❌ Page 'Affectation du résultat' non trouvée.")
        return donnees, 0
    
    # Extraire le tableau
    tables = pdf.pages[page_index].extract_tables()
    if not tables:
        print("   ❌ Aucun tableau extrait.")
        return donnees, 0
    
    table = tables[0]
    print(f"   ✓ Tableau extrait ({len(table)} lignes, {len(table[0]) if table else 0} colonnes)")
    
    # Parcourir toutes les lignes pour trouver les codes
    for row_idx, row in enumerate(table):
        # Chercher les codes dans la ligne
        for col_idx, cell in enumerate(row):
            if not cell:
                continue
            
            code = str(cell).strip().upper()
            
            # AFFECTATION DU RÉSULTAT - Code ZE (Dividendes) → index 26
            if code == "ZE" and code in CODES_AFFECTATION_RESULTAT:
                libelle = CODES_AFFECTATION_RESULTAT[code]
                montant_cell = row[26] if len(row) > 26 else None
                montant = nettoyer_montant(montant_cell)
                
                if montant is not None:
                    donnees.append((libelle, montant))
                    nb_trouves += 1
                    print(f"   ✓ {code} ({libelle}) → {montant} [index 26]")
                else:
                    donnees.append((libelle, 0))
                    print(f"   ⚠️  {code} ({libelle}) → montant non trouvé [index 26]")
            
            # RENSEIGNEMENTS DIVERS - Codes YQ, YR, YT, YU → index 18
            elif code in CODES_RENSEIGNEMENTS_DIVERS:
                libelle = CODES_RENSEIGNEMENTS_DIVERS[code]
                montant_cell = row[18] if len(row) > 18 else None
                montant = nettoyer_montant(montant_cell)
                
                if montant is not None:
                    donnees.append((libelle, montant))
                    nb_trouves += 1
                    print(f"   ✓ {code} ({libelle}) → {montant} [index 18]")
                else:
                    donnees.append((libelle, 0))
                    print(f"   ⚠️  {code} ({libelle}) → montant non trouvé [index 18]")
    
    total_codes = len(CODES_AFFECTATION_RESULTAT) + len(CODES_RENSEIGNEMENTS_DIVERS)
    print(f"   📊 Total : {nb_trouves} valeur(s) trouvée(s) sur {total_codes}")
    return donnees, nb_trouves


def extraire_un_pdf(chemin_pdf):
    """Extrait les données d'un seul PDF.
    
    Returns:
        dict: {'actif': [...], 'passif': [...], 'cr': [...]} ou None en cas d'erreur
    """
    print(f"\n{'='*80}")
    print(f"📄 Traitement : {chemin_pdf.name}")
    print(f"{'='*80}\n")
    
    try:
        with pdfplumber.open(chemin_pdf) as pdf:
            
            # --- ÉTAPE 1 : TROUVER LA PAGE DE L'ACTIF ---
            print("🔍 Recherche de la page du Bilan Actif...")
            actif_page_index = -1
            for i, page in enumerate(pdf.pages):
                text = page.extract_text()
                if text and "Brut" in text and "Net" in text:
                    actif_page_index = i
                    print(f"   ✓ Bilan Actif identifié sur la page {i + 1}.")
                    break
            
            if actif_page_index == -1:
                print("❌ Impossible de trouver la page du Bilan Actif.")
                return None

            # --- ÉTAPE 2 : TROUVER LA PAGE DU PASSIF ---
            print("🔍 Recherche de la page du Bilan Passif...")
            passif_page_index = -1
            for i, page in enumerate(pdf.pages):
                text = page.extract_text()
                if text and "Capital social ou individuel" in text:
                    passif_page_index = i
                    print(f"   ✓ Bilan Passif identifié sur la page {i + 1}.")
                    break
            
            if passif_page_index == -1:
                print("❌ Impossible de trouver la page du Bilan Passif.")
                return None

            # --- ÉTAPE 3 : EXTRAIRE LES TABLEAUX ---
            print("\n📊 Extraction des tableaux...")
            
            tables_actif = pdf.pages[actif_page_index].extract_tables()
            if not tables_actif:
                print(f"❌ Aucun tableau trouvé sur la page de l'Actif.")
                return None
            table_actif = tables_actif[0]
            print(f"   ✓ Tableau Actif extrait.")

            tables_passif = pdf.pages[passif_page_index].extract_tables()
            if not tables_passif:
                print(f"❌ Aucun tableau trouvé sur la page du Passif.")
                return None
            table_passif = tables_passif[0]
            print(f"   ✓ Tableau Passif extrait.")

            # --- ÉTAPE 4 : EXTRACTION DES DONNÉES ---
            
            # Extraction Actif
            print("\n--- 🚀 EXTRACTION DU BILAN ACTIF ---")
            donnees_codes, nb_trouves_codes = extraire_bilan_actif_par_codes(chemin_pdf, table_actif)
            if nb_trouves_codes >= SEUIL_REUSSITE_CODES:
                print("✅ Succès de l'extraction par codes.")
                donnees_actif = donnees_codes
            else:
                print(f"⚠️ Échec par codes ({nb_trouves_codes} valeurs). Basculement sur libellés.")
                donnees_actif = extraire_bilan_actif_par_libelles(chemin_pdf, table_actif)

            # Extraction Passif
            print("\n--- 🚀 EXTRACTION DU BILAN PASSIF ---")
            donnees_codes_passif, nb_trouves_codes_passif = extraire_bilan_passif_par_codes(chemin_pdf, table_passif)
            if nb_trouves_codes_passif >= SEUIL_REUSSITE_CODES_PASSIF:
                print("✅ Succès de l'extraction par codes.")
                donnees_passif = donnees_codes_passif
            else:
                print(f"⚠️ Échec par codes ({nb_trouves_codes_passif} valeurs). Basculement sur libellés.")
                donnees_passif = extraire_bilan_passif_par_libelles(chemin_pdf, table_passif)

            # Extraction Compte de Résultat
            print("\n--- 🚀 EXTRACTION DU COMPTE DE RÉSULTAT ---")
            donnees_codes_cr, nb_trouves_codes_cr = extraire_compte_resultat_par_codes(chemin_pdf, pdf)
            if nb_trouves_codes_cr >= SEUIL_REUSSITE_CODES_COMPTE_RESULTAT:
                print("✅ Succès de l'extraction par codes.")
                donnees_cr = donnees_codes_cr
            else:
                print(f"⚠️ Extraction partielle ({nb_trouves_codes_cr} valeurs).")
                donnees_cr = donnees_codes_cr
            
            # Extraction État des échéances
            print("\n--- 🚀 EXTRACTION DE L'ÉTAT DES ÉCHÉANCES ---")
            donnees_codes_echeances, nb_trouves_codes_echeances = extraire_etat_echeances_par_codes(chemin_pdf, pdf)
            if nb_trouves_codes_echeances >= SEUIL_REUSSITE_CODES_ETAT_ECHEANCES:
                print("✅ Succès de l'extraction par codes.")
                donnees_echeances = donnees_codes_echeances
            else:
                print(f"⚠️ Extraction partielle ({nb_trouves_codes_echeances} valeurs).")
                donnees_echeances = donnees_codes_echeances
            
            # Extraction Affectation du résultat et Renseignements divers
            print("\n--- 🚀 EXTRACTION DE L'AFFECTATION DU RÉSULTAT ET RENSEIGNEMENTS DIVERS ---")
            donnees_codes_affectation, nb_trouves_codes_affectation = extraire_affectation_resultat_par_codes(chemin_pdf, pdf)
            if nb_trouves_codes_affectation >= SEUIL_REUSSITE_CODES_AFFECTATION_RESULTAT:
                print("✅ Succès de l'extraction par codes.")
                donnees_affectation = donnees_codes_affectation
            else:
                print(f"⚠️ Extraction partielle ({nb_trouves_codes_affectation} valeurs).")
                donnees_affectation = donnees_codes_affectation

            return {
                'actif': donnees_actif,
                'passif': donnees_passif,
                'cr': donnees_cr,
                'echeances': donnees_echeances,
                'affectation': donnees_affectation
            }
    
    except Exception as e:
        print(f"❌ Erreur lors du traitement : {e}")
        return None

def main():
    """Version ligne de commande : traite tous les PDFs du dossier 'liasses/' comme une seule année."""
    print("\n" + "="*80)
    print("🚀 EXTRACTION LIASSE FISCALE - MODE CLI")
    print("="*80)
    
    dossier_liasses = Path("liasses")
    dossier_resultats = Path("resultats")
    dossier_resultats.mkdir(exist_ok=True)
    
    fichiers_pdf = list(dossier_liasses.glob("*.pdf"))
    
    if not fichiers_pdf:
        print("\n❌ Aucun PDF dans 'liasses/'\n")
        return
    
    print(f"\n📁 {len(fichiers_pdf)} fichier(s) PDF trouvé(s)")
    print("ℹ️  Mode CLI : Chaque PDF sera traité comme une année différente (2023, 2024, 2025...)")
    print()
    
    donnees_par_annee = {}
    
    # Traiter chaque PDF
    for idx, chemin_pdf in enumerate(fichiers_pdf):
        annee = str(2023 + idx)  # Attribution automatique: 2023, 2024, 2025, etc.
        
        print(f"\n{'='*80}")
        print(f"📄 Fichier {idx + 1}/{len(fichiers_pdf)} : {chemin_pdf.name} → Année {annee}")
        print(f"{'='*80}")
        
        resultats = extraire_un_pdf(chemin_pdf)
        
        if resultats:
            donnees_par_annee[annee] = resultats
            print(f"\n✅ Extraction réussie pour {chemin_pdf.name}")
        else:
            print(f"\n❌ Échec de l'extraction pour {chemin_pdf.name}")
    
    # Générer le fichier Excel
    if donnees_par_annee:
        print(f"\n{'='*80}")
        print("📊 GÉNÉRATION DU FICHIER EXCEL")
        print(f"{'='*80}\n")
        
        nom_excel = dossier_resultats / "extraction_multi_annees.xlsx"
        creer_fichier_excel(donnees_par_annee, nom_excel)
        
        print("="*80)
        print("✅ EXTRACTION TERMINÉE")
        print("="*80)
        print(f"\n📥 Fichier généré : {nom_excel}")
        print(f"📅 Années extraites : {', '.join(sorted(donnees_par_annee.keys()))}")
        print()
    else:
        print("\n❌ Aucune donnée n'a pu être extraite.\n")

if __name__ == "__main__":
    main()
